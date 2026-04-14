"""
pictureBOM — Core library for exporting isometric JPG images of every part
in a SolidWorks assembly and generating an Excel visual BOM with embedded
thumbnails.

This module provides reusable functions with no CLI or GUI side effects.
Use cli.py for the command-line interface or app.py for the web GUI.
"""

import csv
from datetime import datetime
import logging
import os
import re
import time

import pythoncom
import win32com.client
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

log = logging.getLogger(__name__)

# SolidWorks constants
SW_DOC_PART = 1
SW_DOC_ASSEMBLY = 2
SW_OPEN_DOC_OPTIONS_SILENT = 1
SW_VIEW_ISOMETRIC = 7


class PictureBOMError(Exception):
    """Raised when the pipeline encounters a fatal error."""


def connect_to_solidworks():
    """Attach to a running SolidWorks instance."""
    try:
        sw_app = win32com.client.GetActiveObject("SldWorks.Application")
    except pythoncom.com_error:
        raise PictureBOMError(
            "SolidWorks is not running. Please open SolidWorks and try again."
        )
    return sw_app


def open_document(sw_app, file_path, doc_type):
    """Open a SolidWorks document silently. Returns the IModelDoc2 or None."""
    errors = win32com.client.VARIANT(pythoncom.VT_BYREF | pythoncom.VT_I4, 0)
    warnings = win32com.client.VARIANT(pythoncom.VT_BYREF | pythoncom.VT_I4, 0)

    model_doc = sw_app.OpenDoc6(
        file_path,
        doc_type,
        SW_OPEN_DOC_OPTIONS_SILENT,
        "",       # default configuration
        errors,
        warnings,
    )
    return model_doc


def close_document(sw_app, model_doc):
    """Close a document without saving."""
    sw_app.CloseDoc(model_doc.GetTitle)


def sanitize_filename(name):
    """Remove or replace characters that are invalid in file names."""
    return re.sub(r'[<>:"/\\|?*]', "_", name).strip()


def get_component_base_name(component_name):
    """Strip the instance suffix (e.g., 'Bolt-2' -> 'Bolt')."""
    return re.sub(r"-\d+$", "", component_name)


def get_custom_property(cpm, prop_name, debug=False):
    """Read a custom property value. Returns empty string if not found.

    Uses Get6 with all parameters wrapped in explicit VARIANT objects,
    which is required for late-binding COM (no type library).
    Falls back to GetAll3 (bulk read) if Get6 fails.
    """
    # --- Approach 1: Get6 with fully typed VARIANTs ---
    try:
        in_field_name = win32com.client.VARIANT(pythoncom.VT_BSTR, prop_name)
        in_use_cached = win32com.client.VARIANT(pythoncom.VT_I4, 0)  # False as int

        out_val = win32com.client.VARIANT(pythoncom.VT_BYREF | pythoncom.VT_BSTR, None)
        out_resolved = win32com.client.VARIANT(pythoncom.VT_BYREF | pythoncom.VT_BSTR, None)
        out_was_resolved = win32com.client.VARIANT(pythoncom.VT_BYREF | pythoncom.VT_BOOL, None)
        out_link = win32com.client.VARIANT(pythoncom.VT_BYREF | pythoncom.VT_BOOL, None)

        result = cpm.Get6(
            in_field_name,
            in_use_cached,
            out_val,
            out_resolved,
            out_was_resolved,
            out_link,
        )
        if debug:
            log.debug(
                "Get6('%s'): result=%s, val='%s', resolved='%s'",
                prop_name, result, out_val.value, out_resolved.value,
            )
        return str(out_resolved.value or out_val.value or "")
    except Exception as e:
        if debug:
            log.debug("Get6('%s') FAILED: %s, trying GetAll3 fallback...", prop_name, e)

    # --- Approach 2: GetAll3 bulk read fallback ---
    try:
        out_names = win32com.client.VARIANT(pythoncom.VT_VARIANT | pythoncom.VT_BYREF, [])
        out_types = win32com.client.VARIANT(pythoncom.VT_VARIANT | pythoncom.VT_BYREF, [])
        out_values = win32com.client.VARIANT(pythoncom.VT_VARIANT | pythoncom.VT_BYREF, [])
        out_resolved = win32com.client.VARIANT(pythoncom.VT_VARIANT | pythoncom.VT_BYREF, [])
        out_link = win32com.client.VARIANT(pythoncom.VT_VARIANT | pythoncom.VT_BYREF, [])

        cpm.GetAll3(out_names, out_types, out_values, out_resolved, out_link)

        names = out_names.value if out_names.value else ()
        values = out_values.value if out_values.value else ()
        for i, name in enumerate(names):
            if str(name).lower() == prop_name.lower():
                val = str(values[i]) if i < len(values) else ""
                if debug:
                    log.debug("GetAll3 match '%s': '%s'", prop_name, val)
                return val
    except Exception as e2:
        if debug:
            log.debug("GetAll3 fallback FAILED: %s", e2)

    return ""


def get_all_property_names(comp_doc):
    """Get all custom property names from a component. Returns a list of strings."""
    try:
        cpm = comp_doc.Extension.CustomPropertyManager("")
        names = cpm.GetNames
        if names is None:
            return []
        return list(names)
    except Exception:
        return []


def get_part_properties(comp_doc, debug=False):
    """Extract Description, Vendor, and Vendor Part No from a component's custom properties."""
    props = {"description": "", "vendor": "", "vendor_part_no": ""}
    if comp_doc is None:
        return props

    try:
        cpm = comp_doc.Extension.CustomPropertyManager("")

        if debug:
            names = get_all_property_names(comp_doc)
            log.debug("Properties found: %s", names)

        # Build a case-insensitive lookup of actual property names
        all_names = get_all_property_names(comp_doc)
        name_map = {n.lower(): n for n in all_names}

        # Match our target properties case-insensitively
        for target, key in [("description", "description"), ("vendor", "vendor"), ("vendor part no", "vendor_part_no")]:
            actual_name = name_map.get(target)
            if actual_name:
                props[key] = get_custom_property(cpm, actual_name, debug=debug)
    except Exception:
        pass

    return props


def traverse_assembly(assembly_doc, include_subassemblies=False, debug=False):
    """
    Walk the assembly component tree and return a dict of unique components
    with quantities and custom properties (flat list).

    Returns:
        dict: {normalized_file_path: {name, file_path, doc_type, quantity, description, vendor, vendor_part_no}}
    """
    components = assembly_doc.GetComponents(False)  # False = all levels, not just top
    if components is None:
        return {}

    unique = {}

    for comp in components:
        if comp.IsSuppressed:
            continue

        file_path = comp.GetPathName
        if not file_path:
            continue

        normalized = file_path.lower().strip()
        is_assembly = normalized.endswith(".sldasm")

        if is_assembly and not include_subassemblies:
            continue

        if normalized in unique:
            unique[normalized]["quantity"] += 1
            continue

        doc_type = SW_DOC_ASSEMBLY if is_assembly else SW_DOC_PART
        # Use the actual filename (without extension) as the part name
        base_name = os.path.splitext(os.path.basename(file_path))[0]

        # Read custom properties from the component's model doc
        comp_doc = comp.GetModelDoc2
        props = get_part_properties(comp_doc, debug=debug)

        unique[normalized] = {
            "name": base_name,
            "file_path": file_path,
            "doc_type": doc_type,
            "quantity": 1,
            "description": props["description"],
            "vendor": props["vendor"],
            "vendor_part_no": props["vendor_part_no"],
        }

    return unique


def traverse_assembly_hierarchical(assembly_doc, debug=False):
    """
    Walk the assembly tree preserving hierarchy. Returns an ordered list of
    rows with level numbering, type, and per-parent quantities.

    Returns:
        tuple: (rows, unique_components)
            rows: list of dicts with keys: level, type, name, file_path, doc_type,
                  quantity, description, vendor, vendor_part_no
            unique_components: dict keyed by normalized path (for image capture)
    """
    rows = []
    unique = {}

    top_components = assembly_doc.GetComponents(True)  # True = top-level only
    if top_components is None:
        return rows, unique

    _traverse_level(top_components, "", rows, unique, debug)
    return rows, unique


def _traverse_level(components, parent_prefix, rows, unique, debug):
    """Recursively traverse a list of sibling components at one level."""
    # Group siblings by file path to count per-parent quantity
    seen_at_level = {}  # normalized_path -> {comp, count}
    order = []          # preserve insertion order of unique items

    for comp in components:
        if comp.IsSuppressed:
            continue
        file_path = comp.GetPathName
        if not file_path:
            continue

        normalized = file_path.lower().strip()
        if normalized in seen_at_level:
            seen_at_level[normalized]["count"] += 1
        else:
            seen_at_level[normalized] = {"comp": comp, "count": 1}
            order.append(normalized)

    for idx, normalized in enumerate(order, 1):
        info = seen_at_level[normalized]
        comp = info["comp"]
        file_path = comp.GetPathName
        is_assembly = normalized.endswith(".sldasm")
        doc_type = SW_DOC_ASSEMBLY if is_assembly else SW_DOC_PART
        base_name = os.path.splitext(os.path.basename(file_path))[0]

        # Level numbering: top-level = "1.0", "2.0"; children = "1.1", "1.2"
        if parent_prefix == "":
            level = f"{idx}.0"
        else:
            level = f"{parent_prefix}.{idx}"

        comp_doc = comp.GetModelDoc2
        props = get_part_properties(comp_doc, debug=debug)

        row = {
            "level": level,
            "type": "Assembly" if is_assembly else "Part",
            "name": base_name,
            "file_path": file_path,
            "doc_type": doc_type,
            "quantity": info["count"],
            "description": props["description"],
            "vendor": props["vendor"],
            "vendor_part_no": props["vendor_part_no"],
        }
        rows.append(row)

        # Track unique components for image capture
        if normalized not in unique:
            unique[normalized] = {
                "name": base_name,
                "file_path": file_path,
                "doc_type": doc_type,
                "quantity": info["count"],
                "description": props["description"],
                "vendor": props["vendor"],
                "vendor_part_no": props["vendor_part_no"],
            }

        # Recurse into sub-assembly children
        if is_assembly:
            children = comp.GetChildren
            if children:
                # Level prefix for children: "1" for top-level "1.0", "1.1" for "1.1", etc.
                child_prefix = f"{idx}" if parent_prefix == "" else f"{parent_prefix}.{idx}"
                _traverse_level(children, child_prefix, rows, unique, debug)


def _level_depth(level_str):
    """Return nesting depth from a level string.
    '1.0' -> 1 (top-level), '1.1' -> 2, '1.1.3' -> 3
    """
    parts = level_str.split(".")
    if len(parts) == 2 and parts[1] == "0":
        return 1
    return len(parts)


def _build_flat_from_hierarchical(rows, root_assembly_name="Assembly"):
    """Build a flat parts list from hierarchical rows, computing true total
    quantities (multiplied through the assembly hierarchy) and Where Used strings.

    Returns a list of dicts with keys: name, file_path, doc_type,
    total_quantity, description, vendor, vendor_part_no, where_used
    """
    parts = {}       # normalized_path -> accumulator dict
    part_order = []  # preserve first-seen order
    assy_stack = []  # list of (depth, assy_name, cumulative_multiplier)

    for row in rows:
        depth = _level_depth(row["level"])
        per_parent_qty = row["quantity"]

        # Pop ancestors at same depth or deeper (no longer parents)
        while assy_stack and assy_stack[-1][0] >= depth:
            assy_stack.pop()

        if row["type"] == "Assembly":
            parent_mult = assy_stack[-1][2] if assy_stack else 1
            cumulative = parent_mult * per_parent_qty
            assy_stack.append((depth, row["name"], cumulative))
        else:
            # Part — compute total contribution and record parent
            parent_mult = assy_stack[-1][2] if assy_stack else 1
            total_contribution = per_parent_qty * parent_mult
            parent_name = assy_stack[-1][1] if assy_stack else root_assembly_name

            normalized = row["file_path"].lower().strip()
            if normalized not in parts:
                parts[normalized] = {
                    "name": row["name"],
                    "file_path": row["file_path"],
                    "doc_type": row.get("doc_type", SW_DOC_PART),
                    "total_quantity": 0,
                    "description": row["description"],
                    "vendor": row["vendor"],
                    "vendor_part_no": row["vendor_part_no"],
                    "where_used_map": {},
                }
                part_order.append(normalized)

            parts[normalized]["total_quantity"] += total_contribution
            wu = parts[normalized]["where_used_map"]
            if parent_name not in wu:
                wu[parent_name] = per_parent_qty

    # Build final list with where_used string
    result = []
    for norm_path in part_order:
        data = parts[norm_path]
        wu_parts = [f"{name} ({qty})" for name, qty in data["where_used_map"].items()]
        data["where_used"] = ", ".join(wu_parts)
        del data["where_used_map"]
        result.append(data)

    return result


def capture_component(sw_app, file_path, doc_type, output_path, width, height):
    """Open a component, set isometric view, and export a JPG image."""
    model_doc = open_document(sw_app, file_path, doc_type)
    if model_doc is None:
        return False

    try:
        errors = win32com.client.VARIANT(pythoncom.VT_BYREF | pythoncom.VT_I4, 0)
        sw_app.ActivateDoc2(model_doc.GetTitle, False, errors)

        model_doc.ShowNamedView2("*Isometric", SW_VIEW_ISOMETRIC)
        model_doc.ViewZoomtofit2()

        export_data = win32com.client.VARIANT(pythoncom.VT_DISPATCH, None)
        save_errors = win32com.client.VARIANT(pythoncom.VT_BYREF | pythoncom.VT_I4, 0)
        save_warnings = win32com.client.VARIANT(pythoncom.VT_BYREF | pythoncom.VT_I4, 0)
        success = model_doc.Extension.SaveAs(
            output_path,
            0,              # swSaveAsCurrentVersion
            1,              # swSaveAsOptions_Silent
            export_data,    # Nothing (null dispatch)
            save_errors,
            save_warnings,
        )
        return bool(success)
    finally:
        close_document(sw_app, model_doc)


def capture_all_components(sw_app, components, output_dir, width, height, on_progress=None):
    """
    Capture isometric images for all components.

    on_progress: optional callable(current, total, part_name, success, image_path)
    Returns: (success_count, total)
    """
    total = len(components)
    success_count = 0

    for i, (_, comp) in enumerate(components.items(), 1):
        safe_name = sanitize_filename(comp["name"])
        img_output = os.path.join(output_dir, f"{safe_name}.jpg")

        log.info("[%d/%d] Capturing %s...", i, total, comp["name"])
        t0 = time.time()
        try:
            ok = capture_component(
                sw_app, comp["file_path"], comp["doc_type"],
                img_output, width, height,
            )
            if ok:
                success_count += 1
            else:
                log.warning("Failed to open %s", comp["name"])
        except Exception as e:
            ok = False
            log.warning("Error capturing %s: %s", comp["name"], e)

        component_elapsed = time.time() - t0
        if on_progress:
            on_progress(i, total, comp["name"], ok, img_output,
                        elapsed_seconds=component_elapsed)

    return success_count, total


def load_csv_bom(csv_path):
    """Load a CSV file and return a list of row dicts. Expects a 'Part Number' column."""
    rows = []
    with open(csv_path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append(row)
    return rows, reader.fieldnames


def generate_excel_bom(bom_rows, images_dir, output_path, csv_columns=None,
                       hierarchical=False):
    """
    Generate an Excel BOM with embedded thumbnail images.

    bom_rows: list of dicts with at least 'name' key. May also have
              description, quantity, vendor, vendor_part_no.
              In hierarchical mode, also has 'level' and 'type'.
    images_dir: folder containing JPG images named <part_name>.jpg
    output_path: path to write the .xlsx file
    csv_columns: if provided, use these column names (from CSV import)
    hierarchical: if True, add Level and Type columns, highlight assembly rows
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Visual BOM"

    # --- Styles ---
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    assembly_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    # Columns where text should be left-aligned (description-like)
    left_align_names = {"description", "where used (sub-asm qty)"}

    if csv_columns:
        # CSV mode: Picture + all original CSV columns
        headers = ["Picture"] + list(csv_columns)
        ws.append(headers)

        for row_idx, row_data in enumerate(bom_rows, start=2):
            part_number = row_data.get("Part Number", row_data.get("part_number", ""))
            safe_name = sanitize_filename(part_number)
            img_path = _find_image(images_dir, safe_name)

            ws.row_dimensions[row_idx].height = 45

            if img_path:
                _insert_image(ws, img_path, f"A{row_idx}")

            for col_idx, col_name in enumerate(csv_columns, start=2):
                ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, ""))

    elif hierarchical:
        # Hierarchical BOM with Level and Type columns
        headers = ["Picture", "Level", "Type", "Part Number", "Description",
                    "Qty", "Vendor", "Vendor Part No"]
        ws.append(headers)

        for row_idx, row_data in enumerate(bom_rows, start=2):
            safe_name = sanitize_filename(row_data["name"])
            img_path = _find_image(images_dir, safe_name)

            ws.row_dimensions[row_idx].height = 45

            if img_path:
                _insert_image(ws, img_path, f"A{row_idx}")

            ws.cell(row=row_idx, column=2, value=row_data.get("level", ""))
            ws.cell(row=row_idx, column=3, value=row_data.get("type", ""))
            ws.cell(row=row_idx, column=4, value=row_data["name"])
            ws.cell(row=row_idx, column=5, value=row_data.get("description", ""))
            ws.cell(row=row_idx, column=6, value=row_data.get("quantity", 1))
            ws.cell(row=row_idx, column=7, value=row_data.get("vendor", ""))
            ws.cell(row=row_idx, column=8, value=row_data.get("vendor_part_no", ""))

    else:
        # Flat BOM (parts only)
        headers = ["Picture", "Part Number", "Description", "Total Qty",
                    "Vendor", "Vendor Part No", "Where Used (Sub-Asm Qty)"]
        ws.append(headers)

        for row_idx, row_data in enumerate(bom_rows, start=2):
            safe_name = sanitize_filename(row_data["name"])
            img_path = _find_image(images_dir, safe_name)

            ws.row_dimensions[row_idx].height = 45

            if img_path:
                _insert_image(ws, img_path, f"A{row_idx}")

            ws.cell(row=row_idx, column=2, value=row_data["name"])
            ws.cell(row=row_idx, column=3, value=row_data.get("description", ""))
            ws.cell(row=row_idx, column=4, value=row_data.get("total_quantity",
                                                               row_data.get("quantity", 1)))
            ws.cell(row=row_idx, column=5, value=row_data.get("vendor", ""))
            ws.cell(row=row_idx, column=6, value=row_data.get("vendor_part_no", ""))
            ws.cell(row=row_idx, column=7, value=row_data.get("where_used", ""))

    # --- Auto-fit column widths based on content ---
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(headers[col_idx - 1]))  # start with header length
        for row_idx in range(2, len(bom_rows) + 2):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        # Add padding, cap at 60 so descriptions don't stretch forever
        width = min(max_len + 4, 60)
        # Picture column: fixed width for the thumbnail
        if col_idx == 1:
            width = 18
        ws.column_dimensions[col_letter].width = width

    # --- Format header row ---
    ws.row_dimensions[1].height = 30
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # --- Format data rows ---
    num_cols = len(headers)
    for row_idx in range(2, len(bom_rows) + 2):
        row_data = bom_rows[row_idx - 2]
        is_assembly_row = hierarchical and row_data.get("type") == "Assembly"

        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            col_name = headers[col_idx - 1] if col_idx <= len(headers) else ""
            if col_name.lower() in left_align_names:
                cell.alignment = left_align
            else:
                cell.alignment = center_align

            # Highlight assembly rows with a light blue tint
            if is_assembly_row:
                cell.fill = assembly_fill
                if col_idx > 1:  # don't bold the picture cell
                    cell.font = Font(bold=True)

    # Freeze the header row so it stays visible when scrolling
    ws.freeze_panes = "A2"

    wb.save(output_path)
    return output_path


def _format_sheet(ws, headers, num_data_rows, bom_rows, hierarchical=False):
    """Apply shared formatting to a BOM worksheet: header styles, borders,
    column widths, assembly highlighting, and freeze panes."""
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    assembly_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    left_align_names = {"description", "where used (sub-asm qty)"}

    # Auto-fit column widths
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(headers[col_idx - 1]))
        for row_idx in range(2, num_data_rows + 2):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None and not str(val).startswith("="):
                max_len = max(max_len, len(str(val)))
        width = min(max_len + 4, 60)
        if col_idx == 1:
            width = 18
        ws.column_dimensions[col_letter].width = width

    # Format header row
    ws.row_dimensions[1].height = 30
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Format data rows
    num_cols = len(headers)
    for row_idx in range(2, num_data_rows + 2):
        row_data = bom_rows[row_idx - 2] if bom_rows else {}
        is_assembly_row = hierarchical and row_data.get("type") == "Assembly"

        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            col_name = headers[col_idx - 1] if col_idx <= len(headers) else ""
            if col_name.lower() in left_align_names:
                cell.alignment = left_align
            else:
                cell.alignment = center_align

            if is_assembly_row:
                cell.fill = assembly_fill
                if col_idx > 1:
                    cell.font = Font(bold=True)

    ws.freeze_panes = "A2"


def _generate_linked_excel_bom(flat_parts, hierarchical_rows, images_dir, output_path):
    """Generate a two-sheet Excel BOM with XLOOKUP formulas linking the
    Assemblies sheet back to the Parts sheet.

    flat_parts: list of dicts from _build_flat_from_hierarchical()
    hierarchical_rows: list of dicts from traverse_assembly_hierarchical()
    """
    wb = Workbook()

    # ---- Sheet 1: Parts Only (Editable) ----
    ws1 = wb.active
    ws1.title = "Parts Only (Editable)"

    headers1 = ["Picture", "Part Number", "Description", "Total Qty",
                 "Vendor", "Vendor Part No", "Where Used (Sub-Asm Qty)"]
    ws1.append(headers1)

    for row_idx, part in enumerate(flat_parts, start=2):
        safe_name = sanitize_filename(part["name"])
        img_path = _find_image(images_dir, safe_name)

        ws1.row_dimensions[row_idx].height = 45

        if img_path:
            _insert_image(ws1, img_path, f"A{row_idx}")

        ws1.cell(row=row_idx, column=2, value=part["name"])
        ws1.cell(row=row_idx, column=3, value=part.get("description", ""))
        ws1.cell(row=row_idx, column=4, value=part.get("total_quantity", 1))
        ws1.cell(row=row_idx, column=5, value=part.get("vendor", ""))
        ws1.cell(row=row_idx, column=6, value=part.get("vendor_part_no", ""))
        ws1.cell(row=row_idx, column=7, value=part.get("where_used", ""))

    _format_sheet(ws1, headers1, len(flat_parts), flat_parts)

    # ---- Sheet 2: Assemblies (Read-Only) ----
    ws2 = wb.create_sheet("Assemblies (Read-Only)")

    headers2 = ["Picture", "Level", "Type", "Part Number", "Description",
                 "Qty", "Vendor", "Vendor Part No"]
    ws2.append(headers2)

    S1 = "Parts Only (Editable)"  # sheet name for formula references
    last_row = len(flat_parts) + 201  # bounded range + 200 row buffer for user additions

    for row_idx, row_data in enumerate(hierarchical_rows, start=2):
        safe_name = sanitize_filename(row_data["name"])
        img_path = _find_image(images_dir, safe_name)

        ws2.row_dimensions[row_idx].height = 45

        if img_path:
            _insert_image(ws2, img_path, f"A{row_idx}")

        ws2.cell(row=row_idx, column=2, value=row_data.get("level", ""))
        ws2.cell(row=row_idx, column=3, value=row_data.get("type", ""))
        ws2.cell(row=row_idx, column=4, value=row_data["name"])
        ws2.cell(row=row_idx, column=6, value=row_data.get("quantity", 1))

        if row_data.get("type") == "Part":
            # INDEX/MATCH formulas: look up Part Number (col D) in Sheet 1
            B = f"'{S1}'!$B$2:$B${last_row}"  # lookup range (Part Number)
            ws2.cell(row=row_idx, column=5).value = (
                f"=IFERROR(INDEX('{S1}'!$C$2:$C${last_row},MATCH(D{row_idx},{B},0)),\"\")")
            ws2.cell(row=row_idx, column=7).value = (
                f"=IFERROR(INDEX('{S1}'!$E$2:$E${last_row},MATCH(D{row_idx},{B},0)),\"\")")
            ws2.cell(row=row_idx, column=8).value = (
                f"=IFERROR(INDEX('{S1}'!$F$2:$F${last_row},MATCH(D{row_idx},{B},0)),\"\")")
        else:
            # Assembly rows: static values (assemblies aren't on the flat sheet)
            ws2.cell(row=row_idx, column=5, value=row_data.get("description", ""))
            ws2.cell(row=row_idx, column=7, value=row_data.get("vendor", ""))
            ws2.cell(row=row_idx, column=8, value=row_data.get("vendor_part_no", ""))

    # For formula columns on Sheet 2, estimate widths from Sheet 1 data
    _format_sheet(ws2, headers2, len(hierarchical_rows), hierarchical_rows,
                  hierarchical=True)

    # Widen formula columns using Sheet 1 data as estimate
    for s2_col, s1_col in [(5, 3), (7, 5), (8, 6)]:
        letter = get_column_letter(s2_col)
        s1_letter = get_column_letter(s1_col)
        ws2.column_dimensions[letter].width = ws1.column_dimensions[s1_letter].width

    wb.save(output_path)
    return output_path


def _find_image(images_dir, safe_name):
    """Find an image file matching the part name (try .jpg then .bmp)."""
    for ext in (".jpg", ".jpeg", ".bmp", ".png"):
        path = os.path.join(images_dir, safe_name + ext)
        if os.path.isfile(path):
            return path
    return None


def _insert_image(ws, img_path, cell_ref):
    """Insert and size an image into a worksheet cell."""
    img = XlImage(img_path)
    # Scale to fit within ~55px tall (row height 45pt ≈ 60px)
    thumb_height = 55
    aspect = img.width / img.height if img.height else 1.78
    img.height = thumb_height
    img.width = int(thumb_height * aspect)
    img.anchor = cell_ref
    ws.add_image(img)


# ---------------------------------------------------------------------------
# BOM Comparison
# ---------------------------------------------------------------------------


def parse_bom_excel(excel_path):
    """Read a pictureBOM-generated .xlsx and return part data.

    Handles flat, hierarchical, linked (two-sheet), and CSV-mode workbooks
    by reading headers dynamically.

    Returns:
        dict with keys:
            parts: {part_number: {"qty": int, "description": str}}
            images_dir: directory containing the Excel file (where images live)
    """
    wb = load_workbook(excel_path, data_only=True)

    # Pick the best sheet: prefer "Parts Only" (linked mode has total qty)
    ws = wb.active
    for name in wb.sheetnames:
        if name.lower().startswith("parts only"):
            ws = wb[name]
            break

    # Read headers from row 1
    headers = {}  # col_index -> lowercase header name
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value:
            headers[col_idx] = str(cell.value).strip().lower()

    # Find key columns
    pn_col = None
    qty_col = None
    desc_col = None
    level_col = None
    type_col = None

    for col_idx, name in headers.items():
        if name == "part number":
            pn_col = col_idx
        elif name == "total qty":
            qty_col = col_idx
        elif name == "qty" and qty_col is None:
            qty_col = col_idx
        elif name == "description":
            desc_col = col_idx
        elif name == "level":
            level_col = col_idx
        elif name == "type":
            type_col = col_idx

    if pn_col is None:
        raise PictureBOMError(f"No 'Part Number' column found in {excel_path}")

    is_hierarchical = (level_col is not None and type_col is not None
                       and "total qty" not in headers.values())

    if is_hierarchical:
        return _parse_hierarchical_bom(ws, pn_col, qty_col, desc_col,
                                       level_col, type_col, excel_path)

    # Flat / linked / CSV — read part number and qty directly
    parts = {}
    for row in ws.iter_rows(min_row=2):
        pn_val = row[pn_col - 1].value
        if not pn_val:
            continue
        part_number = str(pn_val).strip()
        qty = row[qty_col - 1].value if qty_col else 1
        qty = int(qty) if qty else 1
        desc = str(row[desc_col - 1].value or "") if desc_col else ""

        if part_number in parts:
            parts[part_number]["qty"] += qty
        else:
            parts[part_number] = {"qty": qty, "description": desc}

    wb.close()
    return {"parts": parts, "images_dir": os.path.dirname(os.path.abspath(excel_path))}


def _parse_hierarchical_bom(ws, pn_col, qty_col, desc_col, level_col, type_col,
                            excel_path):
    """Parse a hierarchical BOM sheet, multiplying quantities through the
    assembly tree to compute true totals (same logic as _build_flat_from_hierarchical)."""
    parts = {}
    assy_stack = []  # list of (depth, cumulative_multiplier)

    for row in ws.iter_rows(min_row=2):
        level_val = row[level_col - 1].value
        type_val = row[type_col - 1].value
        pn_val = row[pn_col - 1].value
        if not pn_val or not level_val:
            continue

        part_number = str(pn_val).strip()
        row_type = str(type_val or "").strip()
        per_parent_qty = int(row[qty_col - 1].value or 1) if qty_col else 1
        desc = str(row[desc_col - 1].value or "") if desc_col else ""
        depth = _level_depth(str(level_val))

        # Pop ancestors at same depth or deeper
        while assy_stack and assy_stack[-1][0] >= depth:
            assy_stack.pop()

        if row_type == "Assembly":
            parent_mult = assy_stack[-1][1] if assy_stack else 1
            cumulative = parent_mult * per_parent_qty
            assy_stack.append((depth, cumulative))
        else:
            # Part — compute total contribution
            parent_mult = assy_stack[-1][1] if assy_stack else 1
            total_contribution = per_parent_qty * parent_mult

            if part_number in parts:
                parts[part_number]["qty"] += total_contribution
            else:
                parts[part_number] = {"qty": total_contribution, "description": desc}

    return {"parts": parts, "images_dir": os.path.dirname(os.path.abspath(excel_path))}


def compare_boms(bom_a_path, bom_b_path):
    """Compare two BOM files: find parts in B not fully covered by A.

    Args:
        bom_a_path: Excel file for parts the user already has.
        bom_b_path: Excel file for the assembly the user wants to build.

    Returns:
        dict with keys:
            rows: list of dicts (part_number, description, qty_a, qty_b,
                  shortage, image_path)
            summary: {total_in_b, shortage_count, fully_covered}
            bom_a: basename of file A
            bom_b: basename of file B
    """
    parsed_a = parse_bom_excel(bom_a_path)
    parsed_b = parse_bom_excel(bom_b_path)
    parts_a = parsed_a["parts"]
    parts_b = parsed_b["parts"]

    rows = []
    for part_number in sorted(parts_b.keys()):
        info_b = parts_b[part_number]
        qty_b = info_b["qty"]
        qty_a = parts_a[part_number]["qty"] if part_number in parts_a else 0

        if qty_a >= qty_b:
            continue  # fully covered

        shortage = qty_b - qty_a

        # Find image: prefer B's directory, fall back to A's
        safe_name = sanitize_filename(part_number)
        image_path = _find_image(parsed_b["images_dir"], safe_name)
        if not image_path:
            image_path = _find_image(parsed_a["images_dir"], safe_name)

        rows.append({
            "part_number": part_number,
            "description": info_b["description"],
            "qty_a": qty_a,
            "qty_b": qty_b,
            "shortage": shortage,
            "image_path": image_path,
        })

    total_in_b = len(parts_b)
    shortage_count = len(rows)

    return {
        "rows": rows,
        "summary": {
            "total_in_b": total_in_b,
            "shortage_count": shortage_count,
            "fully_covered": total_in_b - shortage_count,
        },
        "bom_a": os.path.basename(bom_a_path),
        "bom_b": os.path.basename(bom_b_path),
    }


def generate_comparison_excel(comparison, output_path):
    """Write a comparison result to a formatted Excel file with images.

    Args:
        comparison: dict returned by compare_boms()
        output_path: path to write the .xlsx file

    Returns:
        output_path
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Parts to Order"

    rows = comparison["rows"]
    summary = comparison["summary"]

    # Summary row (merged across all columns)
    summary_text = (
        f"Comparing \"{comparison['bom_b']}\" against \"{comparison['bom_a']}\": "
        f"{summary['shortage_count']} part(s) to order, "
        f"{summary['fully_covered']} already covered"
    )
    ws.merge_cells("A1:F1")
    summary_cell = ws.cell(row=1, column=1, value=summary_text)
    summary_cell.font = Font(bold=True, size=11)
    summary_cell.alignment = Alignment(horizontal="left", vertical="center",
                                       wrap_text=True)
    ws.row_dimensions[1].height = 30

    # Headers in row 2
    headers = ["Picture", "Part Number", "Description", "Already Have",
               "Need", "To Order"]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=2, column=col_idx, value=header)

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F3864", end_color="1F3864",
                              fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center",
                                 wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    missing_fill = PatternFill(start_color="FDDEDE", end_color="FDDEDE",
                               fill_type="solid")
    shortage_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD",
                                fill_type="solid")

    # Format header row
    ws.row_dimensions[2].height = 30
    for cell in ws[2]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data rows starting at row 3
    for row_idx, row_data in enumerate(rows, start=3):
        ws.row_dimensions[row_idx].height = 45

        if row_data["image_path"]:
            _insert_image(ws, row_data["image_path"], f"A{row_idx}")

        ws.cell(row=row_idx, column=2, value=row_data["part_number"])
        ws.cell(row=row_idx, column=3, value=row_data["description"])
        ws.cell(row=row_idx, column=4, value=row_data["qty_a"])
        ws.cell(row=row_idx, column=5, value=row_data["qty_b"])
        ws.cell(row=row_idx, column=6, value=row_data["shortage"])

        # Color code: red if completely missing, yellow if partial shortage
        fill = missing_fill if row_data["qty_a"] == 0 else shortage_fill
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.fill = fill
            if col_idx == 3:  # Description
                cell.alignment = left_align
            else:
                cell.alignment = center_align

    # Column widths
    col_widths = [18, 25, 40, 14, 10, 12]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A3"
    wb.save(output_path)
    return output_path


def run_pipeline(assembly_path, output_dir, width=1920, height=1080,
                 include_subassemblies=False, bom_mode=None, csv_path=None,
                 images_dir=None, debug=False, on_progress=None,
                 on_status=None, overwrite=False):
    """
    Run the full pictureBOM pipeline.

    Args:
        assembly_path: Path to the .sldasm file.
        output_dir: Directory for images and BOM output.
        width: Image export width in pixels.
        height: Image export height in pixels.
        include_subassemblies: Legacy flag; use bom_mode instead.
        bom_mode: "flat", "nested", or "linked". If None, derived from
                  include_subassemblies for backward compatibility.
        csv_path: Optional CSV file for BOM data instead of SolidWorks properties.
        images_dir: Optional folder of existing images (skips capture).
        debug: Enable verbose property logging.
        on_progress: Optional callable(current, total, part_name, success, image_path).
        on_status: Optional callable(message) for stage updates.
        overwrite: If True, overwrite existing files without checking.

    Returns:
        dict with keys: excel_path, images_dir, total_components, captured_count

    Raises:
        PictureBOMError: On fatal errors (file not found, SolidWorks not running, etc.)
    """
    # Resolve bom_mode from new or legacy parameter
    if bom_mode is None:
        bom_mode = "nested" if include_subassemblies else "flat"

    def status(msg):
        log.info(msg)
        if on_status:
            on_status(msg)

    has_csv = csv_path is not None
    has_images = images_dir is not None

    assembly_path = os.path.abspath(assembly_path)
    output_dir = os.path.abspath(output_dir)
    os.makedirs(output_dir, exist_ok=True)

    if not os.path.isfile(assembly_path):
        raise PictureBOMError(f"File not found: {assembly_path}")
    if not assembly_path.lower().endswith(".sldasm"):
        raise PictureBOMError("Input file must be a SolidWorks assembly (.sldasm)")

    # Where images live
    img_dir = os.path.abspath(images_dir) if has_images else output_dir

    # Build Excel filename from assembly name + timestamp
    root_name = os.path.splitext(os.path.basename(assembly_path))[0]
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    excel_name = f"{root_name}_{timestamp}.xlsx"
    excel_path = os.path.join(output_dir, excel_name)
    if not overwrite:
        if not has_images:
            existing = [f for f in os.listdir(output_dir)
                        if f.lower().endswith((".jpg", ".jpeg", ".bmp", ".png"))] if os.path.isdir(output_dir) else []
            if existing:
                raise PictureBOMError(
                    f"Output folder already contains {len(existing)} image(s): {output_dir}"
                )

    status("Connecting to SolidWorks...")
    sw_app = connect_to_solidworks()

    assy_name = os.path.basename(assembly_path)
    status(f"Opening assembly: {assy_name}")
    assy_doc = open_document(sw_app, assembly_path, SW_DOC_ASSEMBLY)
    if assy_doc is None:
        raise PictureBOMError("Failed to open assembly file.")

    status("Traversing assembly components...")
    # All modes use hierarchical traversal (flat/linked need it for Where Used)
    hierarchical_rows, components = traverse_assembly_hierarchical(assy_doc, debug=debug)
    total = len(components)
    status(f"Found {total} unique component(s)")

    # BOM data comes from CSV if provided, otherwise from SolidWorks traversal
    csv_columns = None
    bom_rows = None
    if has_csv:
        csv_path = os.path.abspath(csv_path)
        if not os.path.isfile(csv_path):
            raise PictureBOMError(f"CSV file not found: {csv_path}")
        status(f"Loading CSV: {os.path.basename(csv_path)}")
        bom_rows, csv_columns = load_csv_bom(csv_path)
        bom_mode = "flat"  # CSV overrides mode
        status(f"Loaded {len(bom_rows)} rows from CSV")

    # Capture images (skip if user provided existing images)
    captured_count = 0
    capture_start = time.time()
    if not has_images and total > 0:
        status(f"Capturing images ({total} components)...")
        captured_count, _ = capture_all_components(
            sw_app, components, output_dir, width, height, on_progress=on_progress,
        )
        status(f"{captured_count}/{total} images captured.")
    elif has_images:
        status(f"Using existing images from: {img_dir}")
    capture_elapsed = time.time() - capture_start

    close_document(sw_app, assy_doc)

    if not hierarchical_rows and not bom_rows:
        log.warning("No BOM data to write.")
        return {
            "excel_path": None,
            "images_dir": img_dir,
            "total_components": total,
            "captured_count": captured_count,
            "timing": {
                "capture_seconds": round(capture_elapsed, 2),
                "excel_seconds": 0,
                "per_component_avg": round(capture_elapsed / total, 2) if total > 0 else 0,
            },
        }

    # Generate Excel BOM
    status("Generating Excel BOM...")
    excel_start = time.time()

    if csv_columns:
        # CSV mode — flat sheet with CSV data
        generate_excel_bom(bom_rows, img_dir, excel_path, csv_columns=csv_columns)
    elif bom_mode == "linked":
        flat_parts = _build_flat_from_hierarchical(hierarchical_rows, root_name)
        _generate_linked_excel_bom(flat_parts, hierarchical_rows, img_dir, excel_path)
    elif bom_mode == "nested":
        generate_excel_bom(hierarchical_rows, img_dir, excel_path, hierarchical=True)
    else:
        # Flat mode — build flat parts from hierarchical data for Where Used
        flat_parts = _build_flat_from_hierarchical(hierarchical_rows, root_name)
        generate_excel_bom(flat_parts, img_dir, excel_path)

    excel_elapsed = time.time() - excel_start
    status(f"Done! BOM saved to: {excel_path}")

    return {
        "excel_path": excel_path,
        "images_dir": img_dir,
        "total_components": total,
        "captured_count": captured_count,
        "timing": {
            "capture_seconds": round(capture_elapsed, 2),
            "excel_seconds": round(excel_elapsed, 2),
            "per_component_avg": round(capture_elapsed / total, 2) if total > 0 else 0,
        },
    }
