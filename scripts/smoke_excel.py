"""Smoke test for the Excel export pipeline — no SolidWorks needed.

Builds every export mode (flat, nested, CSV, linked, comparison) from canned
BOM dicts plus the example Cage Stack Assembly images, then round-trips each
output through parse_bom_excel and compares against the pre-port openpyxl
workbook. Run after any change to the Excel writers:

    uv run python scripts/smoke_excel.py
"""

import sys
import tempfile
from pathlib import Path

from picturebom.core import (
    BUFFER_ROWS,
    FLAT_HEADERS,
    HIERARCHICAL_HEADERS,
    _build_flat_from_hierarchical,
    _generate_linked_excel_bom,
    _vendor_url,
    compare_boms,
    generate_comparison_excel,
    generate_excel_bom,
    load_bom_table,
    parse_bom_excel,
    parse_property_list,
    write_bom_csv,
)

REPO_ROOT = Path(__file__).resolve().parents[1]
IMAGES_DIR = REPO_ROOT / "Example 3d models" / "Cage Stack Assembly" / "BOM"
# Newest generated workbook in the (machine-local) example folder, else the
# committed showcase workbook — same content, always present in a clone.
_workbooks = sorted(IMAGES_DIR.glob("Cage2-sjm_*.xlsx"))
OLD_WORKBOOK = (_workbooks[-1] if _workbooks
                else REPO_ROOT / "docs" / "samples" / "Cage2-sjm_pictureBOM.xlsx")

# Mimics traverse_assembly_hierarchical() output for a small assembly:
# one subassembly (x2) containing two parts, plus two top-level parts.
def _hrow(level, row_type, name, qty, desc, vendor, vendor_pn):
    return {"level": level, "type": row_type, "name": name, "quantity": qty,
            "description": desc, "vendor": vendor, "vendor_part_no": vendor_pn,
            "file_path": f"C:\\fake\\{name}.sldprt"}


HIER_ROWS = [
    _hrow("1.0", "Assembly", "cag subassm 1", 2, "Cage subassembly", "", ""),
    _hrow("1.1", "Part", "CP33_M-Solidworks", 2, "30 mm Cage Plate",
          "Thorlabs", "CP33/M"),
    _hrow("1.2", "Part", "ER3-Solidworks", 4, "Cage Assembly Rod 3in",
          "", "ER3"),
    _hrow("2.0", "Part", "KC1T_M-Solidworks", 1, "Kinematic Mount",
          "Thorlabs", "KC1T/M"),
    _hrow("3.0", "Part", "91290A115", 8, "M3x10 socket head cap screw",
          "McMaster-Carr", "91290A115"),
]
# Subassembly multiplier is 2, so true totals double the per-parent counts.
EXPECTED_TOTALS = {"CP33_M-Solidworks": 4, "ER3-Solidworks": 8,
                   "KC1T_M-Solidworks": 1, "91290A115": 8}

CSV_COLUMNS = ["Part Number", "Description", "Qty", "Vendor", "Vendor Part No"]
CSV_ROWS = [
    {"Part Number": "CP33_M-Solidworks", "Description": "30 mm Cage Plate",
     "Qty": "4", "Vendor": "Thorlabs", "Vendor Part No": "CP33/M"},
    {"Part Number": "91290A115", "Description": "M3x10 SHCS",
     "Qty": "8", "Vendor": "McMaster-Carr", "Vendor Part No": "91290A115"},
    # '=' prefix must survive as text (strings_to_formulas off), and the
    # None vendor (ragged CSV row) must not pollute the vendor dropdown.
    {"Part Number": "EQ-TEST", "Description": "=see drawing 42",
     "Qty": "1", "Vendor": None, "Vendor Part No": None},
]

failures = []


def check(label, condition, detail=""):
    status = "ok" if condition else "FAIL"
    print(f"  [{status}] {label}" + (f" — {detail}" if detail and not condition else ""))
    if not condition:
        failures.append(label)


def main():
    if not OLD_WORKBOOK.is_file():
        sys.exit(f"Missing reference workbook: {OLD_WORKBOOK}")
    out_dir = Path(tempfile.mkdtemp(prefix="picturebom_smoke_"))
    print(f"Outputs: {out_dir}\n")
    images = str(IMAGES_DIR)

    print("_vendor_url:")
    check("thorlabs by vendor",
          _vendor_url("Thorlabs", "CP33/M")
          == "https://www.thorlabs.com/thorproduct.cfm?partnumber=CP33%2FM")
    check("mcmaster by vendor",
          _vendor_url("McMaster-Carr", "91290A115")
          == "https://www.mcmaster.com/91290A115/")
    check("thorlabs by PN shape", _vendor_url("", "ER3") is not None)
    check("mcmaster by PN shape",
          _vendor_url("", "91290A115") == "https://www.mcmaster.com/91290A115/")
    check("unknown vendor -> no link", _vendor_url("Newport", "XYZ-1") is None)
    check("blank PN -> no link", _vendor_url("Thorlabs", "") is None)

    flat_parts = _build_flat_from_hierarchical(HIER_ROWS, "Cage2-sjm")

    print("\nflat mode:")
    flat_path = out_dir / "flat.xlsx"
    generate_excel_bom(flat_parts, images, str(flat_path))
    parsed = parse_bom_excel(str(flat_path))["parts"]
    check("all parts parsed", set(parsed) == set(EXPECTED_TOTALS),
          f"got {sorted(parsed)}")
    check("total quantities",
          {pn: p["qty"] for pn, p in parsed.items()} == EXPECTED_TOTALS,
          f"got { {pn: p['qty'] for pn, p in parsed.items()} }")

    from openpyxl import load_workbook
    wb = load_workbook(str(flat_path))
    ws = wb.active
    check("Status header is last column",
          ws.cell(row=1, column=8).value == "Status")
    check("blank Status cell is formatted (write_blank regression)",
          ws.cell(row=2, column=8).border.left.style == "thin",
          f"got {ws.cell(row=2, column=8).border.left.style!r}")
    wb.close()

    print("\nnested mode:")
    nested_path = out_dir / "nested.xlsx"
    generate_excel_bom(HIER_ROWS, images, str(nested_path), hierarchical=True)
    parsed = parse_bom_excel(str(nested_path))["parts"]
    check("hierarchical totals multiplied",
          {pn: p["qty"] for pn, p in parsed.items()} == EXPECTED_TOTALS,
          f"got { {pn: p['qty'] for pn, p in parsed.items()} }")

    print("\nCSV mode:")
    csv_path = out_dir / "csv.xlsx"
    generate_excel_bom(CSV_ROWS, images, str(csv_path), csv_columns=CSV_COLUMNS)
    parsed = parse_bom_excel(str(csv_path))["parts"]
    check("csv rows parsed",
          {pn: p["qty"] for pn, p in parsed.items()}
          == {"CP33_M-Solidworks": 4, "91290A115": 8, "EQ-TEST": 1},
          f"got { {pn: p['qty'] for pn, p in parsed.items()} }")
    check("'=' description stays text (strings_to_formulas regression)",
          parsed["EQ-TEST"]["description"] == "=see drawing 42",
          f"got {parsed['EQ-TEST']['description']!r}")
    wb = load_workbook(str(csv_path))
    lists = [c.value for c in wb["Lists"]["A"] if c.value]
    check("no 'None' in vendor dropdown (ragged CSV regression)",
          "None" not in lists, f"got {lists}")
    wb.close()

    print("\nlinked mode:")
    linked_path = out_dir / "linked.xlsx"
    _generate_linked_excel_bom(flat_parts, HIER_ROWS, images, str(linked_path))
    parsed = parse_bom_excel(str(linked_path))["parts"]
    check("Parts Only sheet parsed with totals",
          {pn: p["qty"] for pn, p in parsed.items()} == EXPECTED_TOTALS,
          f"got { {pn: p['qty'] for pn, p in parsed.items()} }")

    print("\npart properties (configured extra columns + autofilter):")
    prop_names, rejected = parse_property_list("Process, Finish, Description")
    check("reserved names rejected",
          (prop_names, rejected) == (["Process", "Finish"], ["Description"]),
          f"got {(prop_names, rejected)}")

    # Same assembly, now with a Process/Finish property on some parts.
    prop_values = [("", ""), ("COTS", ""), ("Machined", "Anodize"),
                   ("COTS", ""), ("COTS", "")]
    props_rows = [dict(r, properties={"Process": proc, "Finish": fin})
                  for r, (proc, fin) in zip(HIER_ROWS, prop_values)]
    props_flat = _build_flat_from_hierarchical(props_rows, "Cage2-sjm")
    check("flat build carries properties",
          props_flat[1]["properties"] == {"Process": "Machined",
                                          "Finish": "Anodize"},
          f"got {props_flat[1]}")

    pflat_path = out_dir / "props_flat.xlsx"
    generate_excel_bom(props_flat, images, str(pflat_path),
                       property_names=prop_names)
    wb = load_workbook(str(pflat_path))
    ws = wb.active
    headers = [c.value for c in ws[1]]
    check("flat: property columns sit before Status",
          headers == ["Picture"] + FLAT_HEADERS + ["Process", "Finish",
                                                   "Status"],
          f"got {headers}")
    check("flat: property value lands in its column",
          ws.cell(row=3, column=8).value == "Machined",  # ER3, col H
          f"got {ws.cell(row=3, column=8).value!r}")
    check("flat: autofilter spans all columns and the buffer rows",
          ws.auto_filter.ref == f"A1:J{len(props_flat) + BUFFER_ROWS + 1}",
          f"got {ws.auto_filter.ref!r}")
    wb.close()

    pnested_path = out_dir / "props_nested.xlsx"
    generate_excel_bom(props_rows, images, str(pnested_path),
                       hierarchical=True, property_names=prop_names)
    wb = load_workbook(str(pnested_path))
    ws = wb.active
    headers = [c.value for c in ws[1]]
    check("nested: property columns sit before Status",
          headers == ["Picture"] + HIERARCHICAL_HEADERS + ["Process",
                                                           "Finish", "Status"],
          f"got {headers}")
    check("nested: autofilter stops at the data rows",
          ws.auto_filter.ref == f"A1:K{len(props_rows) + 1}",
          f"got {ws.auto_filter.ref!r}")
    wb.close()

    plinked_path = out_dir / "props_linked.xlsx"
    _generate_linked_excel_bom(props_flat, props_rows, images,
                               str(plinked_path), property_names=prop_names)
    wb = load_workbook(str(plinked_path))  # formulas, not cached values
    ws1, ws2 = wb["Parts Only (Editable)"], wb["Assemblies (Read-Only)"]
    headers1 = [c.value for c in ws1[1]]
    check("linked: Parts sheet property columns sit before Status",
          headers1 == ["Picture"] + FLAT_HEADERS + ["Process", "Finish",
                                                    "Status"],
          f"got {headers1}")
    check("linked: Parts sheet gets the autofilter, Assemblies doesn't",
          ws1.auto_filter.ref is not None and ws2.auto_filter.ref is None,
          f"got {ws1.auto_filter.ref!r} / {ws2.auto_filter.ref!r}")
    status_dvs = [dv for dv in ws1.data_validations.dataValidation
                  if dv.formula1 and "To Order" in str(dv.formula1)]
    check("linked: Status dropdown followed Status to the last column "
          "(hardcoded-column regression)",
          status_dvs and all(str(dv.sqref).startswith("J")
                             for dv in status_dvs),
          f"got {[str(dv.sqref) for dv in status_dvs]}")
    formula = ws2.cell(row=4, column=9).value  # ER3's Process on Sheet 2
    check("linked: Assemblies property cell is a formula against col H",
          isinstance(formula, str) and formula.startswith("=IFERROR(INDEX(")
          and "$H$" in formula, f"got {formula!r}")
    check("linked: assembly rows get static property cells",
          ws2.cell(row=2, column=9).value in ("", None),
          f"got {ws2.cell(row=2, column=9).value!r}")
    wb.close()
    wb = load_workbook(str(plinked_path), data_only=True)
    cached = wb["Assemblies (Read-Only)"].cell(row=4, column=9).value
    check("linked: property formula carries its cached value",
          cached == "Machined", f"got {cached!r}")
    wb.close()

    ptwin = out_dir / "props_flat.csv"
    write_bom_csv(props_flat, str(ptwin), property_names=prop_names)
    rows, cols = load_bom_table(str(ptwin))
    check("CSV twin carries the property columns",
          cols == FLAT_HEADERS + ["Process", "Finish"], f"got {cols}")
    check("CSV twin carries the property values",
          rows[1]["Process"] == "Machined", f"got {rows[1]}")
    prebuilt = out_dir / "props_rebuilt.xlsx"
    generate_excel_bom(rows, images, str(prebuilt), csv_columns=cols)
    wb = load_workbook(str(prebuilt))
    headers = [c.value for c in wb.active[1]]
    check("rebuild from the twin keeps the property columns",
          "Process" in headers and headers[-1] == "Status", f"got {headers}")
    wb.close()

    print("\nrerun sources (CSV twin and the workbook itself):")
    # Both must feed straight back into a run: same columns, same rows, in
    # the same order — that is what makes a rerun need no SolidWorks.
    for label, rows, kwargs, book in [
        ("flat", flat_parts, {}, flat_path),
        ("nested", HIER_ROWS, {"hierarchical": True}, nested_path),
        ("linked", flat_parts, {}, linked_path),
    ]:
        twin = out_dir / f"{label}.csv"
        write_bom_csv(rows, str(twin), **kwargs)
        csv_rows, csv_cols = load_bom_table(str(twin))
        xlsx_rows, xlsx_cols = load_bom_table(str(book))
        # The workbook carries a Status column (empty on a fresh run) that the
        # CSV twin has no data for; everything else must match exactly.
        check(f"{label}: CSV twin and workbook agree on columns",
              csv_cols == [c for c in xlsx_cols if c.lower() != "status"],
              f"{csv_cols} vs {xlsx_cols}")
        check(f"{label}: CSV twin and workbook agree on rows",
              [{k: str(v) for k, v in r.items()} for r in csv_rows]
              == [{k: str(v) for k, v in r.items() if k.lower() != "status"}
                  for r in xlsx_rows],
              f"{csv_rows[:1]} vs {xlsx_rows[:1]}")
        check(f"{label}: the Picture column is never read back",
              "picture" not in {c.lower() for c in csv_cols + xlsx_cols})
        check(f"{label}: quantities come back as numbers",
              all(isinstance(r[c], (int, float))
                  for r in csv_rows
                  for c in csv_cols if c.lower() in {"qty", "total qty"}),
              f"got {csv_rows[0]}")

    # A rebuild from the twin reproduces the same parts and totals.
    rebuilt = out_dir / "rebuilt_from_csv.xlsx"
    rows, cols = load_bom_table(str(out_dir / "flat.csv"))
    generate_excel_bom(rows, images, str(rebuilt), csv_columns=cols)
    parsed = parse_bom_excel(str(rebuilt))["parts"]
    check("rebuild from the CSV twin matches the original totals",
          {pn: p["qty"] for pn, p in parsed.items()} == EXPECTED_TOTALS,
          f"got { {pn: p['qty'] for pn, p in parsed.items()} }")

    # Rerunning from a workbook someone has been marking up keeps that
    # markup, and must not grow a second Status column.
    marked = load_workbook(str(flat_path))
    marked.active.cell(row=2, column=8).value = "Ordered"
    marked_path = out_dir / "marked.xlsx"
    marked.save(str(marked_path))
    marked.close()
    rows, cols = load_bom_table(str(marked_path))
    rebuilt_marked = out_dir / "rebuilt_from_marked.xlsx"
    generate_excel_bom(rows, images, str(rebuilt_marked), csv_columns=cols)
    wb = load_workbook(str(rebuilt_marked))
    headers = [c.value for c in wb.active[1]]
    check("rerun from a marked-up workbook keeps one Status column",
          [h for h in headers].count("Status") == 1, f"got {headers}")
    check("rerun from a marked-up workbook keeps the Status values",
          wb.active.cell(row=2, column=headers.index("Status") + 1).value == "Ordered",
          f"got {wb.active.cell(row=2, column=headers.index('Status') + 1).value!r}")
    wb.close()

    print("\ncompare old (openpyxl) vs new (xlsxwriter), both directions:")
    cmp_ab = compare_boms(str(OLD_WORKBOOK), str(flat_path))
    cmp_ba = compare_boms(str(flat_path), str(OLD_WORKBOOK))
    check("old-vs-new runs", isinstance(cmp_ab["rows"], list))
    check("new-vs-old runs", isinstance(cmp_ba["rows"], list))

    print("\ncomparison export:")
    cmp_path = out_dir / "comparison.xlsx"
    generate_comparison_excel(cmp_ab if cmp_ab["rows"] else cmp_ba, str(cmp_path))
    from openpyxl import load_workbook
    wb = load_workbook(str(cmp_path), data_only=True)
    check("Parts to Order sheet present", "Parts to Order" in wb.sheetnames)
    wb.close()

    print(f"\n{'ALL CHECKS PASSED' if not failures else f'{len(failures)} FAILURE(S): {failures}'}")
    print(f"Workbooks kept for manual inspection in {out_dir}")
    sys.exit(1 if failures else 0)


if __name__ == "__main__":
    main()
