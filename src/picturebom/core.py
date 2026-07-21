"""
pictureBOM — Core library for exporting isometric JPG images of every part
in a SolidWorks assembly and generating an Excel visual BOM with embedded
thumbnails.

This module provides reusable functions with no CLI or GUI side effects.
Use cli.py for the command-line interface or app.py for the web GUI.
"""

import csv
from datetime import datetime
import logging
import os
import re
import time
from urllib.parse import quote

import pythoncom
import win32com.client
import win32com.client.dynamic
import xlsxwriter
from openpyxl import load_workbook

log = logging.getLogger(__name__)

# SolidWorks constants
SW_DOC_PART = 1
SW_DOC_ASSEMBLY = 2
SW_OPEN_DOC_OPTIONS_SILENT = 1
SW_VIEW_ISOMETRIC = 7

# --- Excel export styling ---
HEADER_BG = "#1F3864"
BORDER_COLOR = "#BFBFBF"
ASSEMBLY_BG = "#D6E4F0"
MISSING_BG = "#FDDEDE"
PARTIAL_BG = "#FFF3CD"
LINK_COLOR = "#0563C1"
PICTURE_COL_WIDTH = 18
DATA_ROW_HEIGHT = 45
HEADER_ROW_HEIGHT = 30
MAX_COL_WIDTH = 60
BUFFER_ROWS = 200  # rows past the data that keep dropdowns/colors working

# Columns where text should be left-aligned (description-like)
LEFT_ALIGN_HEADERS = {"description", "where used (sub-asm qty)"}

STATUS_OPTIONS = ["To Order", "Ordered", "Received", "Installed"]
COMMON_VENDORS = ["Thorlabs", "McMaster-Carr", "Newport", "Digi-Key", "Unknown"]

# Lowercase substring of Vendor -> (fill, font color). The cell recolors as
# soon as a matching vendor is picked (Excel can't color the dropdown
# popup items themselves).
VENDOR_HIGHLIGHTS = {
    "mcmaster": ("#FFEB9C", "#9C6500"),
    "thorlabs": ("#FFC7CE", "#9C0006"),
    "unknown": ("#D9D9D9", "#3F3F3F"),
}

# Lowercase substring of Vendor -> product URL template
VENDOR_URL_TEMPLATES = {
    "thorlabs": "https://www.thorlabs.com/thorproduct.cfm?partnumber={pn}",
    "mcmaster": "https://www.mcmaster.com/{pn}/",
}
# Fallbacks when Vendor is blank: PN shapes are distinctive enough to guess.
_MCMASTER_PN_RE = re.compile(r"^\d{4,6}[A-Z]{1,2}\d{1,4}$")  # e.g. 91290A115
_THORLABS_PN_RE = re.compile(r"^[A-Z]{1,4}\d+[A-Z0-9]*(?:[/-][A-Z0-9]+)*$")  # e.g. KC1T/M


class PictureBOMError(Exception):
    """Raised when the pipeline encounters a fatal error."""


def connect_to_solidworks():
    """Attach to a running SolidWorks instance.

    Force late binding by re-wrapping the raw IDispatch with
    win32com.client.dynamic. This bypasses pywin32's auto-generated typelib
    wrappers (gen_py), which break SolidWorks calls in subtle ways:
    byref VARIANT args raise TypeError, derived interfaces (IAssemblyDoc)
    hide base-interface members (GetTitle), and ActivateDoc2 doesn't
    actually switch the active doc — so SaveAs ends up exporting the same
    view repeatedly under different filenames.
    """
    try:
        sw_app = win32com.client.GetActiveObject("SldWorks.Application")
    except pythoncom.com_error:
        raise PictureBOMError(
            "SolidWorks is not running. Please open SolidWorks and try again."
        )
    return win32com.client.dynamic.Dispatch(sw_app._oleobj_)


def open_document(sw_app, file_path, doc_type):
    """Open a SolidWorks document silently. Returns the IModelDoc2 or None."""
    errors = win32com.client.VARIANT(pythoncom.VT_BYREF | pythoncom.VT_I4, 0)
    warnings = win32com.client.VARIANT(pythoncom.VT_BYREF | pythoncom.VT_I4, 0)

    model_doc = sw_app.OpenDoc6(
        file_path,
        doc_type,
        SW_OPEN_DOC_OPTIONS_SILENT,
        "",       # default configuration
        errors,
        warnings,
    )
    return model_doc


def close_document(sw_app, model_doc):
    """Close a document without saving."""
    sw_app.CloseDoc(model_doc.GetTitle)


def activate_document(sw_app, model_doc):
    """Bring a document to the front in SolidWorks."""
    errors = win32com.client.VARIANT(pythoncom.VT_BYREF | pythoncom.VT_I4, 0)
    sw_app.ActivateDoc2(model_doc.GetTitle, False, errors)


def show_completion_popup(message):
    """Show a topmost Windows info box without blocking the caller.

    Used by the GUI so someone watching SolidWorks (not the browser) sees
    that the run finished.
    """
    import ctypes

    def _popup():
        MB_ICONINFORMATION = 0x40
        MB_SETFOREGROUND = 0x10000
        MB_TOPMOST = 0x40000
        ctypes.windll.user32.MessageBoxW(
            0, message, "pictureBOM",
            MB_ICONINFORMATION | MB_SETFOREGROUND | MB_TOPMOST,
        )

    import threading
    threading.Thread(target=_popup, daemon=True).start()


def sanitize_filename(name):
    """Remove or replace characters that are invalid in file names."""
    return re.sub(r'[<>:"/\\|?*]', "_", name).strip()


def get_component_base_name(component_name):
    """Strip the instance suffix (e.g., 'Bolt-2' -> 'Bolt')."""
    return re.sub(r"-\d+$", "", component_name)


def get_custom_property(cpm, prop_name, debug=False):
    """Read a custom property value. Returns empty string if not found.

    Uses Get6 with all parameters wrapped in explicit VARIANT objects,
    which is required for late-binding COM (no type library).
    Falls back to GetAll3 (bulk read) if Get6 fails.
    """
    # --- Approach 1: Get6 with fully typed VARIANTs ---
    try:
        in_field_name = win32com.client.VARIANT(pythoncom.VT_BSTR, prop_name)
        in_use_cached = win32com.client.VARIANT(pythoncom.VT_I4, 0)  # False as int

        out_val = win32com.client.VARIANT(pythoncom.VT_BYREF | pythoncom.VT_BSTR, None)
        out_resolved = win32com.client.VARIANT(pythoncom.VT_BYREF | pythoncom.VT_BSTR, None)
        out_was_resolved = win32com.client.VARIANT(pythoncom.VT_BYREF | pythoncom.VT_BOOL, None)
        out_link = win32com.client.VARIANT(pythoncom.VT_BYREF | pythoncom.VT_BOOL, None)

        result = cpm.Get6(
            in_field_name,
            in_use_cached,
            out_val,
            out_resolved,
            out_was_resolved,
            out_link,
        )
        if debug:
            log.debug(
                "Get6('%s'): result=%s, val='%s', resolved='%s'",
                prop_name, result, out_val.value, out_resolved.value,
            )
        return str(out_resolved.value or out_val.value or "")
    except Exception as e:
        if debug:
            log.debug("Get6('%s') FAILED: %s, trying GetAll3 fallback...", prop_name, e)

    # --- Approach 2: GetAll3 bulk read fallback ---
    try:
        out_names = win32com.client.VARIANT(pythoncom.VT_VARIANT | pythoncom.VT_BYREF, [])
        out_types = win32com.client.VARIANT(pythoncom.VT_VARIANT | pythoncom.VT_BYREF, [])
        out_values = win32com.client.VARIANT(pythoncom.VT_VARIANT | pythoncom.VT_BYREF, [])
        out_resolved = win32com.client.VARIANT(pythoncom.VT_VARIANT | pythoncom.VT_BYREF, [])
        out_link = win32com.client.VARIANT(pythoncom.VT_VARIANT | pythoncom.VT_BYREF, [])

        cpm.GetAll3(out_names, out_types, out_values, out_resolved, out_link)

        names = out_names.value if out_names.value else ()
        values = out_values.value if out_values.value else ()
        for i, name in enumerate(names):
            if str(name).lower() == prop_name.lower():
                val = str(values[i]) if i < len(values) else ""
                if debug:
                    log.debug("GetAll3 match '%s': '%s'", prop_name, val)
                return val
    except Exception as e2:
        if debug:
            log.debug("GetAll3 fallback FAILED: %s", e2)

    return ""


def get_component_color(comp, comp_doc):
    """Assembly-context color override for a component, read over COM, or None.

    Returns the 9-double MaterialPropertyValues tuple
    (R, G, B, ambient, diffuse, specular, shininess, transparency, emission).
    Deliberately reads ONLY the component-level override: it is exact when an
    engineer colors an instance in the assembly. Part-document colors are NOT
    consulted — modern appearances (how most parts get their look) don't show
    up there, and the property returns the misleading default color instead.
    Parts without an override are colored from their capture image in the
    3D-export fallback (see bomdom.sample_part_colors).
    """
    if comp is None:
        return None
    try:
        vals = comp.MaterialPropertyValues
    except Exception:
        return None
    if vals is None:
        return None
    try:
        vals = tuple(float(v) for v in vals)
    except (TypeError, ValueError):
        return None
    if len(vals) >= 3 and all(0.0 <= v <= 1.0 for v in vals[:3]):
        return vals
    return None


def get_all_property_names(comp_doc):
    """Get all custom property names from a component. Returns a list of strings."""
    try:
        cpm = comp_doc.Extension.CustomPropertyManager("")
        names = cpm.GetNames
        if names is None:
            return []
        return list(names)
    except Exception:
        return []


def get_part_properties(comp_doc, debug=False):
    """Extract Description, Vendor, and Vendor Part No from a component's custom properties."""
    props = {"description": "", "vendor": "", "vendor_part_no": ""}
    if comp_doc is None:
        return props

    try:
        cpm = comp_doc.Extension.CustomPropertyManager("")

        if debug:
            names = get_all_property_names(comp_doc)
            log.debug("Properties found: %s", names)

        # Build a case-insensitive lookup of actual property names
        all_names = get_all_property_names(comp_doc)
        name_map = {n.lower(): n for n in all_names}

        # Match our target properties case-insensitively
        for target, key in [("description", "description"), ("vendor", "vendor"), ("vendor part no", "vendor_part_no")]:
            actual_name = name_map.get(target)
            if actual_name:
                props[key] = get_custom_property(cpm, actual_name, debug=debug)
    except Exception:
        pass

    return props


def traverse_assembly(assembly_doc, include_subassemblies=False, debug=False):
    """
    Walk the assembly component tree and return a dict of unique components
    with quantities and custom properties (flat list).

    Returns:
        dict: {normalized_file_path: {name, file_path, doc_type, quantity, description, vendor, vendor_part_no}}
    """
    components = assembly_doc.GetComponents(False)  # False = all levels, not just top
    if components is None:
        return {}

    unique = {}

    for comp in components:
        if comp.IsSuppressed:
            continue

        file_path = comp.GetPathName
        if not file_path:
            continue

        normalized = file_path.lower().strip()
        is_assembly = normalized.endswith(".sldasm")

        if is_assembly and not include_subassemblies:
            continue

        if normalized in unique:
            unique[normalized]["quantity"] += 1
            continue

        doc_type = SW_DOC_ASSEMBLY if is_assembly else SW_DOC_PART
        # Use the actual filename (without extension) as the part name
        base_name = os.path.splitext(os.path.basename(file_path))[0]

        # Read custom properties from the component's model doc
        comp_doc = comp.GetModelDoc2
        props = get_part_properties(comp_doc, debug=debug)

        unique[normalized] = {
            "name": base_name,
            "file_path": file_path,
            "doc_type": doc_type,
            "quantity": 1,
            "description": props["description"],
            "vendor": props["vendor"],
            "vendor_part_no": props["vendor_part_no"],
        }

    return unique


def traverse_assembly_hierarchical(assembly_doc, debug=False):
    """
    Walk the assembly tree preserving hierarchy. Returns an ordered list of
    rows with level numbering, type, and per-parent quantities.

    Returns:
        tuple: (rows, unique_components)
            rows: list of dicts with keys: level, type, name, file_path, doc_type,
                  quantity, description, vendor, vendor_part_no
            unique_components: dict keyed by normalized path (for image capture)
    """
    rows = []
    unique = {}

    top_components = assembly_doc.GetComponents(True)  # True = top-level only
    if top_components is None:
        return rows, unique

    _traverse_level(top_components, "", rows, unique, debug)
    return rows, unique


def _traverse_level(components, parent_prefix, rows, unique, debug):
    """Recursively traverse a list of sibling components at one level."""
    # Group siblings by file path to count per-parent quantity
    seen_at_level = {}  # normalized_path -> {comp, count}
    order = []          # preserve insertion order of unique items

    for comp in components:
        if comp.IsSuppressed:
            continue
        file_path = comp.GetPathName
        if not file_path:
            continue

        normalized = file_path.lower().strip()
        if normalized in seen_at_level:
            seen_at_level[normalized]["count"] += 1
        else:
            seen_at_level[normalized] = {"comp": comp, "count": 1}
            order.append(normalized)

    for idx, normalized in enumerate(order, 1):
        info = seen_at_level[normalized]
        comp = info["comp"]
        file_path = comp.GetPathName
        is_assembly = normalized.endswith(".sldasm")
        doc_type = SW_DOC_ASSEMBLY if is_assembly else SW_DOC_PART
        base_name = os.path.splitext(os.path.basename(file_path))[0]

        # Level numbering: top-level = "1.0", "2.0"; children = "1.1", "1.2"
        if parent_prefix == "":
            level = f"{idx}.0"
        else:
            level = f"{parent_prefix}.{idx}"

        comp_doc = comp.GetModelDoc2
        props = get_part_properties(comp_doc, debug=debug)
        color = get_component_color(comp, comp_doc)

        row = {
            "level": level,
            "type": "Assembly" if is_assembly else "Part",
            "name": base_name,
            "file_path": file_path,
            "doc_type": doc_type,
            "quantity": info["count"],
            "description": props["description"],
            "vendor": props["vendor"],
            "vendor_part_no": props["vendor_part_no"],
        }
        rows.append(row)

        # Track unique components for image capture
        if normalized not in unique:
            unique[normalized] = {
                "name": base_name,
                "file_path": file_path,
                "doc_type": doc_type,
                "quantity": info["count"],
                "description": props["description"],
                "vendor": props["vendor"],
                "vendor_part_no": props["vendor_part_no"],
                "color": color,
            }

        # Recurse into sub-assembly children
        if is_assembly:
            children = comp.GetChildren
            if children:
                # Level prefix for children: "1" for top-level "1.0", "1.1" for "1.1", etc.
                child_prefix = f"{idx}" if parent_prefix == "" else f"{parent_prefix}.{idx}"
                _traverse_level(children, child_prefix, rows, unique, debug)


def _level_depth(level_str):
    """Return nesting depth from a level string.
    '1.0' -> 1 (top-level), '1.1' -> 2, '1.1.3' -> 3
    """
    parts = level_str.split(".")
    if len(parts) == 2 and parts[1] == "0":
        return 1
    return len(parts)


def _build_flat_from_hierarchical(rows, root_assembly_name="Assembly"):
    """Build a flat parts list from hierarchical rows, computing true total
    quantities (multiplied through the assembly hierarchy) and Where Used strings.

    Returns a list of dicts with keys: name, file_path, doc_type,
    total_quantity, description, vendor, vendor_part_no, where_used
    """
    parts = {}       # normalized_path -> accumulator dict
    part_order = []  # preserve first-seen order
    assy_stack = []  # list of (depth, assy_name, cumulative_multiplier)

    for row in rows:
        depth = _level_depth(row["level"])
        per_parent_qty = row["quantity"]

        # Pop ancestors at same depth or deeper (no longer parents)
        while assy_stack and assy_stack[-1][0] >= depth:
            assy_stack.pop()

        if row["type"] == "Assembly":
            parent_mult = assy_stack[-1][2] if assy_stack else 1
            cumulative = parent_mult * per_parent_qty
            assy_stack.append((depth, row["name"], cumulative))
        else:
            # Part — compute total contribution and record parent
            parent_mult = assy_stack[-1][2] if assy_stack else 1
            total_contribution = per_parent_qty * parent_mult
            parent_name = assy_stack[-1][1] if assy_stack else root_assembly_name

            normalized = row["file_path"].lower().strip()
            if normalized not in parts:
                parts[normalized] = {
                    "name": row["name"],
                    "file_path": row["file_path"],
                    "doc_type": row.get("doc_type", SW_DOC_PART),
                    "total_quantity": 0,
                    "description": row["description"],
                    "vendor": row["vendor"],
                    "vendor_part_no": row["vendor_part_no"],
                    "where_used_map": {},
                }
                part_order.append(normalized)

            parts[normalized]["total_quantity"] += total_contribution
            wu = parts[normalized]["where_used_map"]
            if parent_name not in wu:
                wu[parent_name] = per_parent_qty

    # Build final list with where_used string
    result = []
    for norm_path in part_order:
        data = parts[norm_path]
        wu_parts = [f"{name} ({qty})" for name, qty in data["where_used_map"].items()]
        data["where_used"] = ", ".join(wu_parts)
        del data["where_used_map"]
        result.append(data)

    return result


def capture_component(sw_app, file_path, doc_type, output_path, width, height):
    """Open a component, set isometric view, and export a JPG image."""
    model_doc = open_document(sw_app, file_path, doc_type)
    if model_doc is None:
        return False

    try:
        errors = win32com.client.VARIANT(pythoncom.VT_BYREF | pythoncom.VT_I4, 0)
        sw_app.ActivateDoc2(model_doc.GetTitle, False, errors)

        model_doc.ShowNamedView2("*Isometric", SW_VIEW_ISOMETRIC)
        model_doc.ViewZoomtofit2()

        export_data = win32com.client.VARIANT(pythoncom.VT_DISPATCH, None)
        save_errors = win32com.client.VARIANT(pythoncom.VT_BYREF | pythoncom.VT_I4, 0)
        save_warnings = win32com.client.VARIANT(pythoncom.VT_BYREF | pythoncom.VT_I4, 0)
        success = model_doc.Extension.SaveAs(
            output_path,
            0,              # swSaveAsCurrentVersion
            1,              # swSaveAsOptions_Silent
            export_data,    # Nothing (null dispatch)
            save_errors,
            save_warnings,
        )
        return bool(success)
    finally:
        close_document(sw_app, model_doc)


def capture_all_components(sw_app, components, output_dir, width, height, on_progress=None):
    """
    Capture isometric images for all components.

    on_progress: optional callable(current, total, part_name, success, image_path)
    Returns: (success_count, total)
    """
    total = len(components)
    success_count = 0

    for i, (_, comp) in enumerate(components.items(), 1):
        safe_name = sanitize_filename(comp["name"])
        img_output = os.path.join(output_dir, f"{safe_name}.jpg")

        log.info("[%d/%d] Capturing %s...", i, total, comp["name"])
        t0 = time.time()
        try:
            ok = capture_component(
                sw_app, comp["file_path"], comp["doc_type"],
                img_output, width, height,
            )
            if ok:
                success_count += 1
            else:
                log.warning("Failed to open %s", comp["name"])
        except Exception as e:
            ok = False
            log.warning("Error capturing %s: %s", comp["name"], e)

        component_elapsed = time.time() - t0
        if on_progress:
            on_progress(i, total, comp["name"], ok, img_output,
                        elapsed_seconds=component_elapsed)

    return success_count, total


def export_assembly_glb(sw_app, assy_doc, glb_path):
    """SaveAs the open assembly to a .glb (SolidWorks 2024+ Extended Reality export).

    The exporter acts on the ACTIVE document, and ActivateDoc2 has a history of
    silently failing (see connect_to_solidworks docstring), so activation is
    verified by title before saving. Returns (ok, error_message).
    """
    try:
        title = assy_doc.GetTitle
        active = sw_app.ActiveDoc
        if active is None or active.GetTitle != title:
            activate_document(sw_app, assy_doc)
            active = sw_app.ActiveDoc
            if active is None or active.GetTitle != title:
                return False, "could not bring the assembly to the front for 3D export"

        export_data = win32com.client.VARIANT(pythoncom.VT_DISPATCH, None)
        save_errors = win32com.client.VARIANT(pythoncom.VT_BYREF | pythoncom.VT_I4, 0)
        save_warnings = win32com.client.VARIANT(pythoncom.VT_BYREF | pythoncom.VT_I4, 0)
        success = assy_doc.Extension.SaveAs(
            glb_path,
            0,              # swSaveAsCurrentVersion
            1,              # swSaveAsOptions_Silent
            export_data,
            save_errors,
            save_warnings,
        )
        # SaveAs can return True yet write nothing on versions without the
        # exporter — trust only a real glTF file on disk.
        if not success:
            return False, f"SolidWorks refused the .glb export (error code {save_errors.value})"
        if not os.path.isfile(glb_path) or os.path.getsize(glb_path) <= 12:
            return False, "SolidWorks wrote no 3D data (.glb file missing or empty)"
        with open(glb_path, "rb") as f:
            if f.read(4) != b"glTF":
                return False, "SolidWorks produced an invalid .glb file"
        return True, ""
    except Exception as e:
        return False, f"3D export failed: {e}"


def get_solidworks_year(sw_app):
    """Best-effort SolidWorks release year (e.g. 2024) from RevisionNumber, or None."""
    try:
        major = int(str(sw_app.RevisionNumber).split(".")[0])
        return major + 1992  # SW2024 reports revision 32
    except Exception:
        return None


def load_csv_bom(csv_path):
    """Load a CSV file and return a list of row dicts. Expects a 'Part Number' column."""
    rows = []
    with open(csv_path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append(row)
    return rows, reader.fieldnames


def _create_workbook(output_path):
    """Create an xlsxwriter workbook plus the shared format dictionary.

    xlsxwriter formats are fixed at write time (no post-hoc restyling), so
    every fill/alignment combination the builders need is created up front.
    """
    wb = xlsxwriter.Workbook(str(output_path), {
        "strings_to_formulas": False,  # user text starting with '=' stays text
        "strings_to_urls": False,      # links only via explicit write_url
    })
    border = {"border": 1, "border_color": BORDER_COLOR, "valign": "vcenter"}

    def fmt(**props):
        return wb.add_format({**border, **props})

    fmts = {
        "header": fmt(bold=True, font_color="#FFFFFF", font_size=11,
                      bg_color=HEADER_BG, align="center", text_wrap=True),
        "center": fmt(align="center"),
        "left": fmt(align="left", text_wrap=True),
        "link": fmt(align="center", font_color=LINK_COLOR, underline=1),
        "asm_center": fmt(align="center", bg_color=ASSEMBLY_BG, bold=True),
        "asm_left": fmt(align="left", text_wrap=True, bg_color=ASSEMBLY_BG,
                        bold=True),
        "asm_picture": fmt(align="center", bg_color=ASSEMBLY_BG),  # no bold
        "summary": wb.add_format({"bold": True, "font_size": 11,
                                  "align": "left", "valign": "vcenter",
                                  "text_wrap": True}),
        "missing_center": fmt(align="center", bg_color=MISSING_BG),
        "missing_left": fmt(align="left", text_wrap=True, bg_color=MISSING_BG),
        "partial_center": fmt(align="center", bg_color=PARTIAL_BG),
        "partial_left": fmt(align="left", text_wrap=True, bg_color=PARTIAL_BG),
        # Conditional-format formats carry fill + font only; borders and
        # alignment stay with the underlying cell format.
        "cf": {key: wb.add_format({"bg_color": bg, "font_color": fg})
               for key, (bg, fg) in VENDOR_HIGHLIGHTS.items()},
    }
    return wb, fmts


def _embed_image(ws, row, col, img_path, cell_format):
    """Insert a picture as the cell's value (Excel 365 "Place in Cell").

    In-cell pictures sort/filter with their row and rescale automatically
    when the row height or column width changes. They render in Excel 365
    (2023+); older Excel shows #VALUE! in the cell.
    """
    ws.embed_image(row, col, img_path, {"cell_format": cell_format})


def _vendor_url(vendor, part_no):
    """Return a product-page URL for a part, or None.

    The Vendor property wins when present; a blank vendor falls back to
    part-number shape (McMaster PNs are digit-led, Thorlabs letter-led).
    A wrong guess only costs a dead link, so the patterns stay conservative.
    """
    part_no = str(part_no or "").strip()
    if not part_no:
        return None
    pn_quoted = quote(part_no, safe="")  # Thorlabs "/M" suffix -> %2FM
    vendor_lc = str(vendor or "").strip().lower()
    for key, template in VENDOR_URL_TEMPLATES.items():
        if key in vendor_lc:
            return template.format(pn=pn_quoted)
    if not vendor_lc:
        pn_upper = part_no.upper()
        if _MCMASTER_PN_RE.match(pn_upper):
            return VENDOR_URL_TEMPLATES["mcmaster"].format(pn=pn_quoted)
        if _THORLABS_PN_RE.match(pn_upper):
            return VENDOR_URL_TEMPLATES["thorlabs"].format(pn=pn_quoted)
    return None


def _write_part_link(ws, row, col, url, text, link_fmt, fallback_fmt):
    """Write a hyperlinked part number. write_url writes NOTHING when it
    fails (e.g. URL past Excel's ~2079-char limit), so fall back to plain
    text rather than losing the cell."""
    if ws.write_url(row, col, url, link_fmt, string=text) < 0:
        ws.write(row, col, text, fallback_fmt)


def _vendor_options(discovered):
    """Vendor dropdown options: vendors seen in the BOM, then common vendors
    not already present, with 'Unknown' always last."""
    seen = {}
    for vendor in discovered:
        vendor = str(vendor or "").strip()
        if vendor and vendor.lower() not in seen and vendor.lower() != "unknown":
            seen[vendor.lower()] = vendor
    options = sorted(seen.values(), key=str.lower)
    options += [v for v in COMMON_VENDORS
                if v != "Unknown" and v.lower() not in seen]
    options.append("Unknown")
    return options


def _add_vendor_list_sheet(wb, options):
    """Write dropdown options to a hidden 'Lists' sheet and return the range
    reference for data validation. A sheet-backed list (vs. inline) avoids
    Excel's 255-char inline limit and survives commas in vendor names."""
    ws = wb.add_worksheet("Lists")
    for idx, option in enumerate(options):
        ws.write(idx, 0, option)
    ws.hide()
    return f"='Lists'!$A$1:$A${len(options)}"


def _add_row_extras(ws, first_row, last_row, vendor_col, status_col,
                    vendor_source, cf_fmts):
    """Attach the Status dropdown, Vendor dropdown, and vendor color rules
    to a 0-based row range. The range may extend past the data (buffer
    rows): validation and conditional formatting leave empty cells empty,
    so readers that skip blank rows are unaffected."""
    if status_col is not None:
        ws.data_validation(first_row, status_col, last_row, status_col, {
            "validate": "list", "source": STATUS_OPTIONS,
            "ignore_blank": True})
    if vendor_col is not None:
        if vendor_source:
            ws.data_validation(first_row, vendor_col, last_row, vendor_col, {
                "validate": "list", "source": vendor_source,
                "ignore_blank": True})
        # Excel's text-contains rule is SEARCH-based, i.e. case-insensitive.
        for key, cf in cf_fmts.items():
            ws.conditional_format(first_row, vendor_col, last_row, vendor_col, {
                "type": "text", "criteria": "containing", "value": key,
                "format": cf})


def _part_row_runs(bom_rows):
    """Contiguous 0-based index runs of non-Assembly rows (hierarchical
    mode). Extras apply to parts only; assembly rows are structure."""
    runs = []
    start = None
    for idx, row in enumerate(bom_rows):
        if row.get("type") != "Assembly":
            if start is None:
                start = idx
        elif start is not None:
            runs.append((start, idx - 1))
            start = None
    if start is not None:
        runs.append((start, len(bom_rows) - 1))
    return runs


class _ColWidths:
    """Column-width accumulator: xlsxwriter cells can't be read back after
    writing, so widths are tracked as values are written."""

    def __init__(self, headers):
        self._max = [len(str(h)) for h in headers]

    def track(self, col, value):
        if value is not None:
            self._max[col] = max(self._max[col], len(str(value)))

    def apply(self, ws, picture_col=0):
        for col, max_len in enumerate(self._max):
            if col == picture_col:
                ws.set_column(col, col, PICTURE_COL_WIDTH)
            else:
                # Add padding, cap so descriptions don't stretch forever
                ws.set_column(col, col, min(max_len + 4, MAX_COL_WIDTH))


def _write_flat_rows(ws, fmts, widths, flat_parts, images_dir):
    """Write flat-layout data rows (shared by flat mode and the linked
    Parts sheet). Returns the set of vendor names seen."""
    vendors = set()
    for idx, part in enumerate(flat_parts):
        r = idx + 1
        ws.set_row(r, DATA_ROW_HEIGHT)

        img_path = _find_image(images_dir, sanitize_filename(part["name"]))
        if img_path:
            _embed_image(ws, r, 0, img_path, fmts["center"])
        else:
            ws.write_blank(r, 0, None, fmts["center"])

        vendor = part.get("vendor", "")
        vendors.add(vendor)
        vendor_pn = part.get("vendor_part_no", "")
        cells = [
            (1, part["name"], fmts["center"]),
            (2, part.get("description", ""), fmts["left"]),
            (3, part.get("total_quantity", part.get("quantity", 1)),
             fmts["center"]),
            (4, vendor, fmts["center"]),
            (6, part.get("where_used", ""), fmts["left"]),
        ]
        for col, value, cell_fmt in cells:
            ws.write(r, col, value, cell_fmt)
            widths.track(col, value)

        url = _vendor_url(vendor, vendor_pn)
        if url:
            _write_part_link(ws, r, 5, url, vendor_pn, fmts["link"],
                             fmts["center"])
        else:
            ws.write(r, 5, vendor_pn, fmts["center"])
        widths.track(5, vendor_pn)

        ws.write_blank(r, 7, None, fmts["center"])  # Status
    return vendors


def generate_excel_bom(bom_rows, images_dir, output_path, csv_columns=None,
                       hierarchical=False):
    """
    Generate an Excel BOM with in-cell thumbnail images (Excel 365
    "Place in Cell": pictures sort/filter with their row and rescale with
    row height). Every mode gains a Status dropdown column; sheets with a
    Vendor column also get a vendor dropdown, vendor color highlighting,
    and Thorlabs/McMaster product links on Vendor Part No.

    bom_rows: list of dicts with at least 'name' key. May also have
              description, quantity, vendor, vendor_part_no.
              In hierarchical mode, also has 'level' and 'type'.
    images_dir: folder containing JPG images named <part_name>.jpg
    output_path: path to write the .xlsx file
    csv_columns: if provided, use these column names (from CSV import)
    hierarchical: if True, add Level and Type columns, highlight assembly rows
    """
    wb, fmts = _create_workbook(output_path)
    ws = wb.add_worksheet("Visual BOM")

    if csv_columns:
        headers = ["Picture"] + [str(c) for c in csv_columns] + ["Status"]
    elif hierarchical:
        headers = ["Picture", "Level", "Type", "Part Number", "Description",
                   "Qty", "Vendor", "Vendor Part No", "Status"]
    else:
        headers = ["Picture", "Part Number", "Description", "Total Qty",
                   "Vendor", "Vendor Part No", "Where Used (Sub-Asm Qty)",
                   "Status"]
    status_col = len(headers) - 1

    widths = _ColWidths(headers)
    ws.set_row(0, HEADER_ROW_HEIGHT)
    for col, header in enumerate(headers):
        ws.write(0, col, header, fmts["header"])

    if csv_columns:
        # CSV mode: Picture + all original CSV columns (+ Status)
        vendor_col = vendor_pn_col = None
        for col_idx, name in enumerate(csv_columns, start=1):
            name_lc = str(name).strip().lower()
            if name_lc == "vendor":
                vendor_col = col_idx
            elif name_lc in {"vendor part no", "vendor part number",
                             "vendor pn"}:
                vendor_pn_col = col_idx

        vendors = set()
        for idx, row_data in enumerate(bom_rows):
            r = idx + 1
            ws.set_row(r, DATA_ROW_HEIGHT)
            part_number = row_data.get("Part Number",
                                       row_data.get("part_number", ""))
            img_path = _find_image(images_dir, sanitize_filename(part_number))
            if img_path:
                _embed_image(ws, r, 0, img_path, fmts["center"])
            else:
                ws.write_blank(r, 0, None, fmts["center"])

            # `or ""` guards ragged CSV rows: DictReader fills missing
            # trailing fields with None, and str(None) is "None".
            vendor = (str(row_data.get(csv_columns[vendor_col - 1]) or "")
                      if vendor_col else "")
            vendors.add(vendor)
            for col_idx, col_name in enumerate(csv_columns, start=1):
                value = row_data.get(col_name, "")
                cell_fmt = (fmts["left"]
                            if str(col_name).strip().lower() in LEFT_ALIGN_HEADERS
                            else fmts["center"])
                url = (_vendor_url(vendor, value)
                       if col_idx == vendor_pn_col else None)
                if url:
                    _write_part_link(ws, r, col_idx, url, str(value),
                                     fmts["link"], cell_fmt)
                else:
                    ws.write(r, col_idx, value, cell_fmt)
                widths.track(col_idx, value)
            ws.write_blank(r, status_col, None, fmts["center"])

        vendor_source = (_add_vendor_list_sheet(wb, _vendor_options(vendors))
                         if vendor_col else None)
        _add_row_extras(ws, 1, len(bom_rows) + BUFFER_ROWS, vendor_col,
                        status_col, vendor_source, fmts["cf"])

    elif hierarchical:
        vendor_col = 6
        vendors = set()
        for idx, row_data in enumerate(bom_rows):
            r = idx + 1
            ws.set_row(r, DATA_ROW_HEIGHT)
            is_assembly = row_data.get("type") == "Assembly"
            center = fmts["asm_center"] if is_assembly else fmts["center"]
            left = fmts["asm_left"] if is_assembly else fmts["left"]

            img_path = _find_image(images_dir,
                                   sanitize_filename(row_data["name"]))
            pic_fmt = fmts["asm_picture"] if is_assembly else fmts["center"]
            if img_path:
                _embed_image(ws, r, 0, img_path, pic_fmt)
            else:
                ws.write_blank(r, 0, None, pic_fmt)

            vendor = row_data.get("vendor", "")
            if not is_assembly:
                vendors.add(vendor)
            cells = [
                (1, row_data.get("level", ""), center),
                (2, row_data.get("type", ""), center),
                (3, row_data["name"], center),
                (4, row_data.get("description", ""), left),
                (5, row_data.get("quantity", 1), center),
                (6, vendor, center),
            ]
            for col, value, cell_fmt in cells:
                ws.write(r, col, value, cell_fmt)
                widths.track(col, value)

            vendor_pn = row_data.get("vendor_part_no", "")
            url = None if is_assembly else _vendor_url(vendor, vendor_pn)
            if url:
                _write_part_link(ws, r, 7, url, vendor_pn, fmts["link"],
                                 center)
            else:
                ws.write(r, 7, vendor_pn, center)
            widths.track(7, vendor_pn)

            ws.write_blank(r, status_col, None, center)

        vendor_source = _add_vendor_list_sheet(wb, _vendor_options(vendors))
        # Extras on part rows only — assembly rows are structure, and the
        # vendor colors would fight the assembly tint.
        for start, end in _part_row_runs(bom_rows):
            _add_row_extras(ws, start + 1, end + 1, vendor_col, status_col,
                            vendor_source, fmts["cf"])

    else:
        vendors = _write_flat_rows(ws, fmts, widths, bom_rows, images_dir)
        vendor_source = _add_vendor_list_sheet(wb, _vendor_options(vendors))
        _add_row_extras(ws, 1, len(bom_rows) + BUFFER_ROWS, 4, status_col,
                        vendor_source, fmts["cf"])

    widths.apply(ws)
    ws.freeze_panes(1, 0)  # keep the header row visible when scrolling
    ws.activate()
    wb.close()
    return output_path


def _generate_linked_excel_bom(flat_parts, hierarchical_rows, images_dir, output_path):
    """Generate a two-sheet Excel BOM with INDEX/MATCH formulas linking the
    Assemblies sheet back to the Parts sheet (INDEX/MATCH kept over XLOOKUP
    for compatibility with older Excel).

    flat_parts: list of dicts from _build_flat_from_hierarchical()
    hierarchical_rows: list of dicts from traverse_assembly_hierarchical()
    """
    wb, fmts = _create_workbook(output_path)

    # ---- Sheet 1: Parts Only (Editable) ----
    # Added first so it stays the workbook's default sheet (parse_bom_excel
    # prefers it). Status must remain the LAST column: the Assemblies sheet
    # references $B/$C/$E/$F by letter.
    ws1 = wb.add_worksheet("Parts Only (Editable)")
    headers1 = ["Picture", "Part Number", "Description", "Total Qty",
                "Vendor", "Vendor Part No", "Where Used (Sub-Asm Qty)",
                "Status"]
    widths1 = _ColWidths(headers1)
    ws1.set_row(0, HEADER_ROW_HEIGHT)
    for col, header in enumerate(headers1):
        ws1.write(0, col, header, fmts["header"])

    vendors = _write_flat_rows(ws1, fmts, widths1, flat_parts, images_dir)

    vendor_source = _add_vendor_list_sheet(wb, _vendor_options(vendors))
    _add_row_extras(ws1, 1, len(flat_parts) + BUFFER_ROWS, 4, 7,
                    vendor_source, fmts["cf"])
    widths1.apply(ws1)
    ws1.freeze_panes(1, 0)

    # ---- Sheet 2: Assemblies (Read-Only) ----
    ws2 = wb.add_worksheet("Assemblies (Read-Only)")
    headers2 = ["Picture", "Level", "Type", "Part Number", "Description",
                "Qty", "Vendor", "Vendor Part No"]
    widths2 = _ColWidths(headers2)
    ws2.set_row(0, HEADER_ROW_HEIGHT)
    for col, header in enumerate(headers2):
        ws2.write(0, col, header, fmts["header"])

    S1 = "Parts Only (Editable)"  # sheet name for formula references
    # Bounded range + buffer for user additions; matches the validation
    # range on Sheet 1 (0-based rows 1..len+BUFFER_ROWS = A2:A{last_row}).
    last_row = len(flat_parts) + BUFFER_ROWS + 1
    # Cached values for formula cells so readers that don't recalculate
    # (openpyxl data_only, file previews) still see real strings.
    parts_by_name = {p["name"]: p for p in flat_parts}

    def write_cell(row, col, value, cell_fmt):
        ws2.write(row, col, value, cell_fmt)
        widths2.track(col, value)

    for idx, row_data in enumerate(hierarchical_rows):
        r = idx + 1
        excel_row = r + 1
        ws2.set_row(r, DATA_ROW_HEIGHT)
        is_assembly = row_data.get("type") == "Assembly"
        center = fmts["asm_center"] if is_assembly else fmts["center"]
        left = fmts["asm_left"] if is_assembly else fmts["left"]

        img_path = _find_image(images_dir, sanitize_filename(row_data["name"]))
        pic_fmt = fmts["asm_picture"] if is_assembly else fmts["center"]
        if img_path:
            _embed_image(ws2, r, 0, img_path, pic_fmt)
        else:
            ws2.write_blank(r, 0, None, pic_fmt)

        write_cell(r, 1, row_data.get("level", ""), center)
        write_cell(r, 2, row_data.get("type", ""), center)
        write_cell(r, 3, row_data["name"], center)
        write_cell(r, 5, row_data.get("quantity", 1), center)

        if not is_assembly:
            # INDEX/MATCH formulas: look up Part Number (col D) in Sheet 1
            part = parts_by_name.get(row_data["name"], row_data)
            B = f"'{S1}'!$B$2:$B${last_row}"  # lookup range (Part Number)
            for col, s1_col, field, cell_fmt in (
                    (4, "C", "description", left),
                    (6, "E", "vendor", center),
                    (7, "F", "vendor_part_no", center)):
                formula = (
                    f"=IFERROR(INDEX('{S1}'!${s1_col}$2:${s1_col}${last_row},"
                    f"MATCH(D{excel_row},{B},0)),\"\")")
                cached = part.get(field, "")
                ws2.write_formula(r, col, formula, cell_fmt, cached)
                widths2.track(col, cached)
        else:
            # Assembly rows: static values (assemblies aren't on the flat sheet)
            write_cell(r, 4, row_data.get("description", ""), left)
            write_cell(r, 6, row_data.get("vendor", ""), center)
            write_cell(r, 7, row_data.get("vendor_part_no", ""), center)

    # Vendor colors on Sheet 2's part rows too (conditional formatting
    # evaluates the formulas' computed values). No dropdowns here.
    for start, end in _part_row_runs(hierarchical_rows):
        _add_row_extras(ws2, start + 1, end + 1, 6, None, None, fmts["cf"])

    widths2.apply(ws2)
    ws2.freeze_panes(1, 0)
    ws1.activate()
    wb.close()
    return output_path


def _find_image(images_dir, safe_name):
    """Find an image file matching the part name (try .jpg then .bmp)."""
    for ext in (".jpg", ".jpeg", ".bmp", ".png"):
        path = os.path.join(images_dir, safe_name + ext)
        if os.path.isfile(path):
            return path
    return None


# ---------------------------------------------------------------------------
# BOM Comparison
# ---------------------------------------------------------------------------


def parse_bom_excel(excel_path):
    """Read a pictureBOM-generated .xlsx and return part data.

    Handles flat, hierarchical, linked (two-sheet), and CSV-mode workbooks
    by reading headers dynamically.

    Returns:
        dict with keys:
            parts: {part_number: {"qty": int, "description": str}}
            images_dir: directory containing the Excel file (where images live)
    """
    wb = load_workbook(excel_path, data_only=True)

    # Pick the best sheet: prefer "Parts Only" (linked mode has total qty)
    ws = wb.active
    for name in wb.sheetnames:
        if name.lower().startswith("parts only"):
            ws = wb[name]
            break

    # Read headers from row 1
    headers = {}  # col_index -> lowercase header name
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value:
            headers[col_idx] = str(cell.value).strip().lower()

    # Find key columns
    pn_col = None
    qty_col = None
    desc_col = None
    level_col = None
    type_col = None

    for col_idx, name in headers.items():
        if name == "part number":
            pn_col = col_idx
        elif name == "total qty":
            qty_col = col_idx
        elif name == "qty" and qty_col is None:
            qty_col = col_idx
        elif name == "description":
            desc_col = col_idx
        elif name == "level":
            level_col = col_idx
        elif name == "type":
            type_col = col_idx

    if pn_col is None:
        raise PictureBOMError(f"No 'Part Number' column found in {excel_path}")

    is_hierarchical = (level_col is not None and type_col is not None
                       and "total qty" not in headers.values())

    if is_hierarchical:
        return _parse_hierarchical_bom(ws, pn_col, qty_col, desc_col,
                                       level_col, type_col, excel_path)

    # Flat / linked / CSV — read part number and qty directly
    parts = {}
    for row in ws.iter_rows(min_row=2):
        pn_val = row[pn_col - 1].value
        if not pn_val:
            continue
        part_number = str(pn_val).strip()
        qty = row[qty_col - 1].value if qty_col else 1
        qty = int(qty) if qty else 1
        desc = str(row[desc_col - 1].value or "") if desc_col else ""

        if part_number in parts:
            parts[part_number]["qty"] += qty
        else:
            parts[part_number] = {"qty": qty, "description": desc}

    wb.close()
    return {"parts": parts, "images_dir": os.path.dirname(os.path.abspath(excel_path))}


def _parse_hierarchical_bom(ws, pn_col, qty_col, desc_col, level_col, type_col,
                            excel_path):
    """Parse a hierarchical BOM sheet, multiplying quantities through the
    assembly tree to compute true totals (same logic as _build_flat_from_hierarchical)."""
    parts = {}
    assy_stack = []  # list of (depth, cumulative_multiplier)

    for row in ws.iter_rows(min_row=2):
        level_val = row[level_col - 1].value
        type_val = row[type_col - 1].value
        pn_val = row[pn_col - 1].value
        if not pn_val or not level_val:
            continue

        part_number = str(pn_val).strip()
        row_type = str(type_val or "").strip()
        per_parent_qty = int(row[qty_col - 1].value or 1) if qty_col else 1
        desc = str(row[desc_col - 1].value or "") if desc_col else ""
        depth = _level_depth(str(level_val))

        # Pop ancestors at same depth or deeper
        while assy_stack and assy_stack[-1][0] >= depth:
            assy_stack.pop()

        if row_type == "Assembly":
            parent_mult = assy_stack[-1][1] if assy_stack else 1
            cumulative = parent_mult * per_parent_qty
            assy_stack.append((depth, cumulative))
        else:
            # Part — compute total contribution
            parent_mult = assy_stack[-1][1] if assy_stack else 1
            total_contribution = per_parent_qty * parent_mult

            if part_number in parts:
                parts[part_number]["qty"] += total_contribution
            else:
                parts[part_number] = {"qty": total_contribution, "description": desc}

    return {"parts": parts, "images_dir": os.path.dirname(os.path.abspath(excel_path))}


def compare_boms(bom_a_path, bom_b_path):
    """Compare two BOM files: find parts in B not fully covered by A.

    Args:
        bom_a_path: Excel file for parts the user already has.
        bom_b_path: Excel file for the assembly the user wants to build.

    Returns:
        dict with keys:
            rows: list of dicts (part_number, description, qty_a, qty_b,
                  shortage, image_path)
            summary: {total_in_b, shortage_count, fully_covered}
            bom_a: basename of file A
            bom_b: basename of file B
    """
    parsed_a = parse_bom_excel(bom_a_path)
    parsed_b = parse_bom_excel(bom_b_path)
    parts_a = parsed_a["parts"]
    parts_b = parsed_b["parts"]

    rows = []
    for part_number in sorted(parts_b.keys()):
        info_b = parts_b[part_number]
        qty_b = info_b["qty"]
        qty_a = parts_a[part_number]["qty"] if part_number in parts_a else 0

        if qty_a >= qty_b:
            continue  # fully covered

        shortage = qty_b - qty_a

        # Find image: prefer B's directory, fall back to A's
        safe_name = sanitize_filename(part_number)
        image_path = _find_image(parsed_b["images_dir"], safe_name)
        if not image_path:
            image_path = _find_image(parsed_a["images_dir"], safe_name)

        rows.append({
            "part_number": part_number,
            "description": info_b["description"],
            "qty_a": qty_a,
            "qty_b": qty_b,
            "shortage": shortage,
            "image_path": image_path,
        })

    total_in_b = len(parts_b)
    shortage_count = len(rows)

    return {
        "rows": rows,
        "summary": {
            "total_in_b": total_in_b,
            "shortage_count": shortage_count,
            "fully_covered": total_in_b - shortage_count,
        },
        "bom_a": os.path.basename(bom_a_path),
        "bom_b": os.path.basename(bom_b_path),
    }


def generate_comparison_excel(comparison, output_path):
    """Write a comparison result to a formatted Excel file with images.

    Args:
        comparison: dict returned by compare_boms()
        output_path: path to write the .xlsx file

    Returns:
        output_path
    """
    wb, fmts = _create_workbook(output_path)
    ws = wb.add_worksheet("Parts to Order")

    rows = comparison["rows"]
    summary = comparison["summary"]

    # Summary row (merged across all columns)
    summary_text = (
        f"Comparing \"{comparison['bom_b']}\" against \"{comparison['bom_a']}\": "
        f"{summary['shortage_count']} part(s) to order, "
        f"{summary['fully_covered']} already covered"
    )
    ws.merge_range(0, 0, 0, 5, summary_text, fmts["summary"])
    ws.set_row(0, HEADER_ROW_HEIGHT)

    # Headers in row 2
    headers = ["Picture", "Part Number", "Description", "Already Have",
               "Need", "To Order"]
    ws.set_row(1, HEADER_ROW_HEIGHT)
    for col, header in enumerate(headers):
        ws.write(1, col, header, fmts["header"])

    # Data rows; color code: red if completely missing, yellow if partial
    for idx, row_data in enumerate(rows):
        r = idx + 2
        ws.set_row(r, DATA_ROW_HEIGHT)
        kind = "missing" if row_data["qty_a"] == 0 else "partial"
        center = fmts[f"{kind}_center"]

        if row_data["image_path"]:
            _embed_image(ws, r, 0, row_data["image_path"], center)
        else:
            ws.write_blank(r, 0, None, center)
        ws.write(r, 1, row_data["part_number"], center)
        ws.write(r, 2, row_data["description"], fmts[f"{kind}_left"])
        ws.write(r, 3, row_data["qty_a"], center)
        ws.write(r, 4, row_data["qty_b"], center)
        ws.write(r, 5, row_data["shortage"], center)

    for col, width in enumerate([PICTURE_COL_WIDTH, 25, 40, 14, 10, 12]):
        ws.set_column(col, col, width)

    ws.freeze_panes(2, 0)
    wb.close()
    return output_path


def run_pipeline(assembly_path, output_dir, width=1920, height=1080,
                 include_subassemblies=False, bom_mode=None, csv_path=None,
                 images_dir=None, debug=False, on_progress=None,
                 on_status=None, overwrite=False, completion_popup=False,
                 output_excel=True, output_html=False,
                 html_size_limit_mb=25, keep_raw_glb=False, viewer_exports=True):
    """
    Run the full pictureBOM pipeline.

    Args:
        assembly_path: Path to the .sldasm file.
        output_dir: Directory for images and BOM output.
        width: Image export width in pixels.
        height: Image export height in pixels.
        include_subassemblies: Legacy flag; use bom_mode instead.
        bom_mode: "flat", "nested", or "linked". If None, derived from
                  include_subassemblies for backward compatibility.
        csv_path: Optional CSV file for BOM data instead of SolidWorks properties.
        images_dir: Optional folder of existing images (skips capture).
        debug: Enable verbose property logging.
        on_progress: Optional callable(current, total, part_name, success, image_path).
        on_status: Optional callable(message) for stage updates.
        overwrite: If True, overwrite existing files without checking.
        completion_popup: If True, show a topmost Windows message box when the
                          BOM is done (for users watching SolidWorks, not the
                          browser).
        output_excel: Write the Excel BOM (default True).
        output_html: Also export an interactive 3D BOM (single .html; needs
                     SolidWorks 2024+ for the .glb export). 3D failures never
                     block the Excel output — they surface as warnings.
        html_size_limit_mb: Above this projected HTML size, geometry is written
                            as a sidecar .glb next to the HTML instead of embedded.
        keep_raw_glb: Keep the intermediate SolidWorks .glb export on disk.
        viewer_exports: Show the Export menu (xlsx/CSV/print) inside the
                        exported HTML viewer. Also hand-editable after export
                        via the "allow_exports" flag near the top of the HTML.

    Returns:
        dict with keys: excel_path, images_dir, total_components, captured_count,
        html_path, html_mode, sidecar_path, html_projected_mb, warnings, timing.

    Raises:
        PictureBOMError: On fatal errors (file not found, SolidWorks not running, etc.)
    """
    # Resolve bom_mode from new or legacy parameter
    if bom_mode is None:
        bom_mode = "nested" if include_subassemblies else "flat"

    def status(msg):
        log.info(msg)
        if on_status:
            on_status(msg)

    has_csv = csv_path is not None
    has_images = images_dir is not None

    assembly_path = os.path.abspath(assembly_path)
    output_dir = os.path.abspath(output_dir)
    os.makedirs(output_dir, exist_ok=True)

    if not os.path.isfile(assembly_path):
        raise PictureBOMError(f"File not found: {assembly_path}")
    if not assembly_path.lower().endswith(".sldasm"):
        raise PictureBOMError("Input file must be a SolidWorks assembly (.sldasm)")

    # Where images live
    img_dir = os.path.abspath(images_dir) if has_images else output_dir

    # Build Excel filename from assembly name + timestamp
    root_name = os.path.splitext(os.path.basename(assembly_path))[0]
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    excel_name = f"{root_name}_{timestamp}.xlsx"
    excel_path = os.path.join(output_dir, excel_name)
    if not overwrite:
        if not has_images:
            existing = [f for f in os.listdir(output_dir)
                        if f.lower().endswith((".jpg", ".jpeg", ".bmp", ".png"))] if os.path.isdir(output_dir) else []
            if existing:
                raise PictureBOMError(
                    f"Output folder already contains {len(existing)} image(s): {output_dir}"
                )

    status("Connecting to SolidWorks...")
    sw_app = connect_to_solidworks()

    if output_html:
        sw_year = get_solidworks_year(sw_app)
        if sw_year is not None and sw_year < 2024:
            status(f"Note: the 3D interactive BOM needs SolidWorks 2024 or newer "
                   f"(this looks like {sw_year}) — will attempt anyway")

    assy_name = os.path.basename(assembly_path)
    status(f"Opening assembly: {assy_name}")
    assy_doc = open_document(sw_app, assembly_path, SW_DOC_ASSEMBLY)
    if assy_doc is None:
        raise PictureBOMError("Failed to open assembly file.")

    status("Traversing assembly components...")
    # All modes use hierarchical traversal (flat/linked need it for Where Used)
    hierarchical_rows, components = traverse_assembly_hierarchical(assy_doc, debug=debug)
    total = len(components)
    status(f"Found {total} unique component(s)")

    # BOM data comes from CSV if provided, otherwise from SolidWorks traversal
    csv_columns = None
    bom_rows = None
    if has_csv:
        csv_path = os.path.abspath(csv_path)
        if not os.path.isfile(csv_path):
            raise PictureBOMError(f"CSV file not found: {csv_path}")
        status(f"Loading CSV: {os.path.basename(csv_path)}")
        bom_rows, csv_columns = load_csv_bom(csv_path)
        bom_mode = "flat"  # CSV overrides mode
        status(f"Loaded {len(bom_rows)} rows from CSV")

    # Capture images (skip if user provided existing images)
    captured_count = 0
    capture_start = time.time()
    if not has_images and total > 0:
        status(f"Capturing images ({total} components)...")
        captured_count, _ = capture_all_components(
            sw_app, components, output_dir, width, height, on_progress=on_progress,
        )
        status(f"{captured_count}/{total} images captured.")
    elif has_images:
        status(f"Using existing images from: {img_dir}")
    capture_elapsed = time.time() - capture_start

    # Leave the assembly open and bring it back to the front — ending on a
    # blank SolidWorks screen reads as "something went wrong". Re-runs are
    # fine: OpenDoc6 returns the already-open document.
    activate_document(sw_app, assy_doc)

    if not hierarchical_rows and not bom_rows:
        log.warning("No BOM data to write.")
        return {
            "excel_path": None,
            "images_dir": img_dir,
            "total_components": total,
            "captured_count": captured_count,
            "html_path": None,
            "html_mode": None,
            "sidecar_path": None,
            "html_projected_mb": None,
            "warnings": [],
            "timing": {
                "capture_seconds": round(capture_elapsed, 2),
                "excel_seconds": 0,
                "glb_seconds": 0,
                "html_seconds": 0,
                "per_component_avg": round(capture_elapsed / total, 2) if total > 0 else 0,
            },
        }

    # Generate Excel BOM
    excel_elapsed = 0.0
    if output_excel:
        status("Generating Excel BOM...")
        excel_start = time.time()

        if csv_columns:
            # CSV mode — flat sheet with CSV data
            generate_excel_bom(bom_rows, img_dir, excel_path, csv_columns=csv_columns)
        elif bom_mode == "linked":
            flat_parts = _build_flat_from_hierarchical(hierarchical_rows, root_name)
            _generate_linked_excel_bom(flat_parts, hierarchical_rows, img_dir, excel_path)
        elif bom_mode == "nested":
            generate_excel_bom(hierarchical_rows, img_dir, excel_path, hierarchical=True)
        else:
            # Flat mode — build flat parts from hierarchical data for Where Used
            flat_parts = _build_flat_from_hierarchical(hierarchical_rows, root_name)
            generate_excel_bom(flat_parts, img_dir, excel_path)

        excel_elapsed = time.time() - excel_start
        status(f"Excel BOM saved to: {excel_path}")
    else:
        excel_path = None

    # Interactive 3D BOM (BomDom). Runs after Excel so a 3D failure can never
    # cost the user their workbook; every failure here degrades to a warning.
    warnings = []
    html_result = {"html_path": None, "html_mode": None, "sidecar_path": None,
                   "html_projected_mb": None}
    glb_elapsed = html_elapsed = 0.0
    if output_html and hierarchical_rows:
        status("Exporting 3D model (single step — may take a few minutes)...")
        glb_start = time.time()
        raw_glb = os.path.join(output_dir, f"{root_name}_{timestamp}_raw.glb")
        glb_ok, glb_err = export_assembly_glb(sw_app, assy_doc, raw_glb)
        glb_elapsed = time.time() - glb_start

        if not glb_ok:
            sw_year = get_solidworks_year(sw_app)
            hint = (f" 3D export needs SolidWorks 2024 or newer (detected {sw_year})."
                    if sw_year is not None and sw_year < 2024 else "")
            msg = f"3D export skipped: {glb_err}.{hint}"
            if output_excel:
                msg += " The Excel BOM was still generated."
            warnings.append(msg)
            status(msg)
        else:
            status(f"3D model exported ({os.path.getsize(raw_glb) / 1e6:.1f} MB)")
            try:
                active_config = ""
                try:
                    active_config = assy_doc.ConfigurationManager.ActiveConfiguration.Name
                except Exception:
                    pass
                try:
                    from importlib.metadata import version as _pkg_version
                    app_version = _pkg_version("picturebom")
                except Exception:
                    app_version = ""

                from . import bomdom  # lazy: keeps module import light

                html_start = time.time()
                res = bomdom.export_bomdom_html(
                    raw_glb, output_dir, root_name, timestamp,
                    hierarchical_rows=hierarchical_rows,
                    flat_parts=_build_flat_from_hierarchical(hierarchical_rows, root_name),
                    bom_names=[c["name"] for c in components.values()],
                    component_colors={c["name"]: c.get("color")
                                      for c in components.values() if c.get("color")},
                    bom_mode=bom_mode,
                    images_dir=img_dir,
                    assembly_file=assy_name,
                    active_config=active_config,
                    app_version=app_version,
                    generated=datetime.now().isoformat(timespec="seconds"),
                    on_status=on_status,
                    size_limit_mb=html_size_limit_mb,
                    viewer_exports=viewer_exports,
                )
                html_elapsed = time.time() - html_start
                warnings.extend(res.pop("warnings", []))
                res.pop("reconciliation", None)
                res.pop("stats", None)
                html_result.update(res)
                status(f"Interactive 3D BOM saved to: {html_result['html_path']}")
                if not keep_raw_glb:
                    try:
                        os.remove(raw_glb)
                    except OSError:
                        pass
            except Exception as e:
                log.exception("BomDom HTML build failed")
                msg = f"3D interactive BOM failed: {e}."
                if output_excel:
                    msg += " The Excel BOM was still generated."
                msg += f" The raw 3D export was kept for diagnosis: {raw_glb}"
                warnings.append(msg)
                status(msg)

    if completion_popup:
        outputs = []
        if excel_path:
            outputs.append("the Excel file")
        if html_result["html_path"]:
            outputs.append("the interactive 3D BOM (.html)")
        popup_msg = ("Your visual BOM is ready.\n\n"
                     "Switch to the pictureBOM tab in your browser to see it, "
                     f"or find {' and '.join(outputs) or 'the output'} in your output folder.")
        if output_html and not html_result["html_path"]:
            popup_msg += "\n\nNote: the 3D export did not complete — see the log."
        show_completion_popup(popup_msg)

    return {
        "excel_path": excel_path,
        "images_dir": img_dir,
        "total_components": total,
        "captured_count": captured_count,
        "html_path": html_result["html_path"],
        "html_mode": html_result["html_mode"],
        "sidecar_path": html_result["sidecar_path"],
        "html_projected_mb": html_result["html_projected_mb"],
        "warnings": warnings,
        "timing": {
            "capture_seconds": round(capture_elapsed, 2),
            "excel_seconds": round(excel_elapsed, 2),
            "glb_seconds": round(glb_elapsed, 2),
            "html_seconds": round(html_elapsed, 2),
            "per_component_avg": round(capture_elapsed / total, 2) if total > 0 else 0,
        },
    }
